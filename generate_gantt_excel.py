#!/usr/bin/env python3
"""
Generador de Gantt VERIFEX v9
- 30 tareas reales del proyecto + 11 version milestones
- Secuencial desde 01/09/2025, entrega 21/08/2026
- Cada tarea tiene un archivo/funcion real en el codigo
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, timedelta

BASE = "/Users/maarco_serrano/Downloads/verifex-standalone 2"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Gantt VERIFEX"

MARCO_FILL = PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")
MARCO_FONT = Font(name="Calibri", size=10, bold=True, color="4A148C")
TONY_FILL = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
TONY_FONT = Font(name="Calibri", size=10, bold=True, color="B71C1C")
LUIS_FILL = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
LUIS_FONT = Font(name="Calibri", size=10, bold=True, color="0D47A1")
ULISES_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
ULISES_FONT = Font(name="Calibri", size=10, bold=True, color="1B5E20")

VERSION_FILL = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
VERSION_FONT = Font(name="Calibri", size=10, bold=True, color="E65100")

header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
phase_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
phase_fill = PatternFill(start_color="3949AB", end_color="3949AB", fill_type="solid")
thin_border = Border(
    left=Side(style="thin", color="BDBDBD"),
    right=Side(style="thin", color="BDBDBD"),
    top=Side(style="thin", color="BDBDBD"),
    bottom=Side(style="thin", color="BDBDBD"),
)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

headers = ["Fase", "Tipo", "#", "Tarea", "Asignado a", "Inicio", "Fin", "Dias", "Depende de", "Estado"]
col_widths = [32, 10, 5, 64, 14, 13, 13, 7, 14, 15]

for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(col_idx)].width = w

# Each entry: (marker, task_num, task_name, assignee, duration_days, depends_on, tipo)
# marker = None (regular task), "V" (version milestone)
# tipo = "Ciclo" (phase header), "Sprint" (regular task), "Iteracion" (milestone)
task_defs = [
    # ═══════════════════════════════════════════
    # CICLO 0: GESTION DE TESIS
    # ═══════════════════════════════════════════
    # ── Fase 0: Eleccion y Preparacion del Tema ──
    (None, None, None, None, None, None, None),

    (None, 1, "Eleccion del tema de tesis: deteccion de noticias falsas con IA", "Todos", 7, None, "Sprint"),
    (None, 2, "Investigacion preliminar y revision bibliografica (fake news, verificacion, IA)", "Todos", 14, 1, "Sprint"),
    (None, 3, "Definicion del problema, preguntas de investigacion y objetivos", "Ulises", 5, 2, "Sprint"),
    (None, 4, "Seleccion de tecnologias: Groq API vs Ollama, React vs Vue, Flask vs FastAPI", "Marco", 5, 3, "Sprint"),
    (None, 5, "Justificacion del proyecto y delimitacion del alcance", "Ulises", 5, 4, "Sprint"),
    (None, 6, "Aprobacion formal del tema de tesis", "Todos", 3, 5, "Sprint"),

    # ═══════════════════════════════════════════
    # CICLO 1: FUNDAMENTOS TEORICOS Y REQUISITOS
    # ═══════════════════════════════════════════
    # ── Fase 1: Marco Teorico e Historias de Usuario ──
    (None, None, None, None, None, None, None),

    (None, 7, "Redaccion del Capitulo 1: Introduccion (contexto, problema, objetivos, justificacion)", "Ulises", 12, 6, "Sprint"),
    (None, 8, "Definicion de 10 historias de usuario (HU-01 a HU-10)", "Marco", 10, 7, "Sprint"),
    (None, 9, "Investigacion sobre IA, Groq API, procesamiento de lenguaje natural", "Marco", 8, 8, "Sprint"),
    (None, 10, "Redaccion del Capitulo 2: Marco Teorico (fake news, IA, Groq API, verificacion)", "Ulises", 14, 9, "Sprint"),
    (None, 11, "Definicion de requisitos funcionales y no funcionales", "Marco", 8, 10, "Sprint"),
    (None, 12, "Creacion de diagramas UML (casos de uso, actividades, secuencia, clases, ER)", "Tony", 12, 11, "Sprint"),

    # ═══════════════════════════════════════════
    # CICLO 2: DISENO DEL SISTEMA
    # ═══════════════════════════════════════════
    # ── Fase 2: Diseno y Metodologia ──
    (None, None, None, None, None, None, None),

    (None, 13, "Redaccion del Capitulo 3: Metodologia y Diseno del Sistema (inicio)", "Ulises", 8, 12, "Sprint"),
    (None, 14, "Diseno de wireframes y mockups de interfaz de usuario", "Luis", 10, 13, "Sprint"),
    (None, 15, "Seleccion de paleta de colores y fuentes (Orbitron, Rajdhani, Share Tech Mono)", "Luis", 5, 14, "Sprint"),
    (None, 16, "Diseno de arquitectura cliente-servidor (React + Flask + Groq API)", "Marco", 8, 15, "Sprint"),
    (None, 17, "Redaccion del Capitulo 3: Finalizacion (arquitectura, componentes, flujo)", "Ulises", 6, 16, "Sprint"),

    # ═══════════════════════════════════════════
    # CICLO 3: DESARROLLO DEL SISTEMA
    # ═══════════════════════════════════════════
    # ── Fase 3: Implementacion ──
    (None, None, None, None, None, None, None),

    (None, 18, "Setup del proyecto (Vite + React + TypeScript + Tailwind + PostCSS)", "Marco", 5, 17, "Sprint"),
    ("V", None, "v0.0.1 - Esqueleto del proyecto con Vite, React, TypeScript y Tailwind configurados", None, 1, 18, "Iteracion"),

    (None, 19, "Setup del backend (Flask + CORS + rutas /analyze y /health)", "Marco", 5, 18, "Sprint"),
    (None, 20, "Implementacion del scraper de URLs con BeautifulSoup (scrape_url)", "Marco", 10, 19, "Sprint"),
    ("V", None, "v0.1.0 - Scraper de URLs funcional (extrae titulo, descripcion y cuerpo)", None, 1, 20, "Iteracion"),

    (None, 21, "Integracion con Groq API (call_groq con fallback a multiples modelos)", "Marco", 10, 20, "Sprint"),
    (None, 22, "Implementacion de prompt engineering (SYSTEM_PROMPT con clasificacion 5 categorias)", "Marco", 8, 21, "Sprint"),
    ("V", None, "v0.2.0 - Analisis con IA via Groq API implementado", None, 1, 22, "Iteracion"),

    (None, 23, "Implementacion del clasificador de credibilidad (analyze_url con 5 categorias)", "Marco", 12, 22, "Sprint"),
    (None, 24, "Implementacion de lista de dominios confiables (CREDIBLE_DOMAINS, 29 fuentes)", "Marco", 4, 23, "Sprint"),
    ("V", None, "v0.3.0 - Clasificador de credibilidad funcional (REAL, FALSO, SATIRA, ESTAFA, NO VERIFICABLE)", None, 1, 24, "Iteracion"),

    (None, 25, "Busqueda de noticias similares via Google News RSS (news_finder.py)", "Marco", 8, 24, "Sprint"),
    (None, 26, "Conexion frontend-backend via API REST (fetch /analyze con AbortController)", "Marco", 8, 25, "Sprint"),
    ("V", None, "v0.4.0 - Frontend conectado al backend via API REST", None, 1, 26, "Iteracion"),

    (None, 27, "Implementacion de componentes base (UrlInput, VerdictDisplay, ConfidenceBar, RedFlags, SimilarNews, LanguageToggle)", "Luis", 14, 26, "Sprint"),
    (None, 28, "Implementacion de manejo de estados (loading, error, results) en App.tsx", "Luis", 8, 27, "Sprint"),
    ("V", None, "v0.5.0 - UI completa con todos los componentes, estados y noticias similares", None, 1, 28, "Iteracion"),

    (None, 29, "Implementacion de soporte bilingue ES/EN en todos los componentes", "Luis", 8, 28, "Sprint"),
    ("V", None, "v0.6.0 - Soporte bilingue ES/EN implementado", None, 1, 29, "Iteracion"),

    (None, 30, "Implementacion de article_type (5 tipos) y deteccion de estafas (is_scam)", "Marco", 8, 29, "Sprint"),
    ("V", None, "v0.7.0 - Clasificacion avanzada: tipo de articulo y deteccion de estafas", None, 1, 30, "Iteracion"),

    (None, 31, "Diseno visual cyberpunk: animacion de cuadricula, glitch, vignette CRT, scanlines, clip-paths", "Luis", 12, 30, "Sprint"),
    ("V", None, "v0.8.0 - Diseno visual cyberpunk finalizado", None, 1, 31, "Iteracion"),

    (None, 32, "Redaccion del Capitulo 4: Desarrollo e Implementacion (concurrente con programacion)", "Ulises", 20, 31, "Sprint"),

    # ═══════════════════════════════════════════
    # CICLO 4: VALIDACION Y DESPLIEGUE
    # ═══════════════════════════════════════════
    # ── Fase 4: Pruebas y Despliegue ──
    (None, None, None, None, None, None, None),

    (None, 33, "Creacion de suite de pruebas backend (server/test_analyzer.py, pytest)", "Marco", 10, 32, "Sprint"),
    (None, 34, "Creacion de suite de pruebas frontend (vitest + testing-library, 7 archivos, 52 tests)", "Marco", 10, 33, "Sprint"),
    (None, 35, "Configuracion de despliegue (Procfile con gunicorn, variables de entorno, CORS)", "Marco", 5, 34, "Sprint"),
    (None, 36, "Despliegue en Railway (fallo por limitaciones, migracion a Render)", "Marco", 7, 35, "Sprint"),
    (None, 37, "Despliegue exitoso en Render con dominio publico", "Marco", 8, 36, "Sprint"),
    ("V", None, "v1.0.0 - Version estable desplegada en Render con dominio publico", None, 1, 37, "Iteracion"),

    (None, 38, "Capturas de pantalla del sistema (interfaz, analisis, resultados, errores, mockups)", "Luis", 5, 37, "Sprint"),

    # ═══════════════════════════════════════════
    # CICLO 5: RESULTADOS Y DOCUMENTACION FINAL
    # ═══════════════════════════════════════════
    # ── Fase 5: Resultados y Documentacion ──
    (None, None, None, None, None, None, None),

    (None, 39, "Pruebas de analisis con URLs reales (positivas, negativas, estafas, satira, opinion)", "Marco", 14, 37, "Sprint"),
    (None, 40, "Correccion de errores: timeouts (AbortController 60s), override dominios, edge cases", "Marco", 12, 39, "Sprint"),
    ("V", None, "v1.1.0 - Refinamiento y correccion de errores del clasificador", None, 1, 40, "Iteracion"),

    (None, 41, "Analisis de casos de prueba y documentacion de resultados", "Tony", 10, 40, "Sprint"),
    (None, 42, "Redaccion del Capitulo 5: Resultados, pruebas y conclusiones", "Ulises", 14, 41, "Sprint"),
    (None, 43, "Redaccion de introduccion, resumen y abstract de la tesis", "Ulises", 5, 42, "Sprint"),
    (None, 44, "Creacion del diagrama de Gantt del proyecto (Gantt_VERIFEX.xlsx)", "Ulises", 4, 43, "Sprint"),
    (None, 45, "Creacion del tablero Kanban del proyecto (Kanban_VERIFEX.xlsx)", "Ulises", 3, 44, "Sprint"),
    (None, 46, "Maquetacion y formato del documento Word (estilos, indices, referencias, portada)", "Ulises", 7, 45, "Sprint"),
    (None, 47, "Revision de contenido, ortografia y consistencia de la tesis", "Tony", 6, 46, "Sprint"),
    (None, 48, "Correcciones finales y ajustes segun retroalimentacion", "Ulises", 5, 47, "Sprint"),
    (None, 49, "Generacion de PDF final de tesis (TESIS_VERIFEX.pdf)", "Ulises", 3, 48, "Sprint"),
    ("V", None, "v1.2.0 - Version final de tesis con todas las funcionalidades implementadas y documentadas", None, 1, 49, "Iteracion"),
]

phase_names = [
    "Fase 0: Eleccion y Preparacion del Tema",
    "Fase 1: Marco Teorico e Historias de Usuario",
    "Fase 2: Diseno y Metodologia",
    "Fase 3: Implementacion",
    "Fase 4: Pruebas y Despliegue",
    "Fase 5: Resultados y Documentacion",
]

cycle_names = [
    "Ciclo 0: Gestion de Tesis",
    "Ciclo 1: Fundamentos Teoricos y Requisitos",
    "Ciclo 2: Diseno del Sistema",
    "Ciclo 3: Desarrollo del Sistema",
    "Ciclo 4: Validacion y Despliegue",
    "Ciclo 5: Resultados y Documentacion Final",
]

assignee_styles = {
    "Marco": (MARCO_FILL, MARCO_FONT),
    "Luis": (LUIS_FILL, LUIS_FONT),
    "Ulises": (ULISES_FILL, ULISES_FONT),
    "Tony": (TONY_FILL, TONY_FONT),
    "Todos": (PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"),
              Font(name="Calibri", size=10, bold=True, color="616161")),
}

status_map = {"H": "Completado", "P": "En Progreso", "N": "Pendiente", "V": "Lanzado"}
status_fills = {
    "H": PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
    "P": PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid"),
    "N": PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid"),
    "V": PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"),
}
status_fonts = {
    "H": Font(name="Calibri", size=10, bold=True, color="2E7D32"),
    "P": Font(name="Calibri", size=10, bold=True, color="F57F17"),
    "N": Font(name="Calibri", size=10, bold=True, color="C62828"),
    "V": Font(name="Calibri", size=10, bold=True, color="E65100"),
}

task_dates = {}
phase_idx = -1
row = 2
version_counter = 0

cycle_fill = PatternFill(start_color="C5CAE9", end_color="C5CAE9", fill_type="solid")
tipo_font = Font(name="Calibri", size=10, bold=True, color="1565C0")

for td in task_defs:
    marker, task_num, task_name, assignee, duration, dep, tipo = td
    ncols = 10

    if marker is None and task_num is None:
        phase_idx += 1

        # Add cycle header before each phase
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        cell = ws.cell(row=row, column=1, value=cycle_names[phase_idx])
        cell.font = Font(name="Calibri", bold=True, size=12, color="1A237E")
        cell.fill = cycle_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = thin_border
        for c in range(2, ncols + 1):
            ws.cell(row=row, column=c).border = thin_border
            ws.cell(row=row, column=c).fill = cycle_fill
        row += 1

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        cell = ws.cell(row=row, column=1, value=phase_names[phase_idx])
        cell.font = phase_font
        cell.fill = phase_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = thin_border
        for c in range(2, ncols + 1):
            ws.cell(row=row, column=c).border = thin_border
            ws.cell(row=row, column=c).fill = phase_fill
        row += 1
        continue

    if marker == "V":
        dep_key = f"T{dep}"
        if dep_key in task_dates:
            start = task_dates[dep_key][1] + timedelta(days=1)
        else:
            start = date(2025, 9, 1)
        end = start + timedelta(days=duration)

        vkey = f"V{version_counter}"
        task_dates[vkey] = (start, end)
        version_counter += 1

        dep_str = str(dep) if dep else ""
        values = ["", tipo, "★", task_name, "—", start, end, duration + 1, dep_str, "Lanzado"]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val if val != "" else None)
            cell.font = VERSION_FONT
            cell.fill = VERSION_FILL
            cell.border = thin_border
            cell.alignment = center if col_idx in (2, 3, 5, 6, 7, 8, 9, 10) else left_wrap
            if col_idx in (6, 7):
                cell.number_format = "DD/MM/YYYY"
    else:
        key = f"T{task_num}"
        dep_key = f"T{dep}" if dep else None
        if dep_key and dep_key in task_dates:
            start = task_dates[dep_key][1] + timedelta(days=1)
        else:
            start = date(2025, 9, 1)
        end = start + timedelta(days=duration)
        task_dates[key] = (start, end)

        if task_num <= 38:
            status = "H"
        elif task_num <= 40:
            status = "P"
        else:
            status = "N"

        status_text = status_map[status]
        a_fill, a_font = assignee_styles.get(assignee, (None, Font(name="Calibri", size=10)))

        dep_str = str(dep) if dep else ""
        values = ["", tipo, task_num, task_name, assignee, start, end, duration + 1, dep_str, status_text]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val if val != "" else None)
            cell.font = Font(name="Calibri", size=10)
            cell.border = thin_border
            cell.alignment = center if col_idx in (2, 3, 5, 6, 7, 8, 9, 10) else left_wrap
            if col_idx in (6, 7):
                cell.number_format = "DD/MM/YYYY"
            elif col_idx == 2:
                cell.font = tipo_font
            elif col_idx == 5:
                if a_fill:
                    cell.fill = a_fill
                cell.font = a_font
            elif col_idx == 10:
                cell.fill = status_fills.get(status)
                cell.font = status_fonts.get(status)
    row += 1

# ── Legend ──
row += 2
ws.cell(row=row, column=1, value="Leyenda:").font = Font(name="Calibri", bold=True, size=11)
row += 1
legend_items = [
    ("Marco - Backend, frontend, IA, deploy y pruebas", MARCO_FILL, MARCO_FONT),
    ("Luis - Frontend, diseno UI/UX y estilos visuales", LUIS_FILL, LUIS_FONT),
    ("Ulises - Documentacion y redaccion de tesis", ULISES_FILL, ULISES_FONT),
    ("Tony - Diagramas UML, documentacion y revision de logica", TONY_FILL, TONY_FONT),
    ("★ Iteracion - Hito de lanzamiento de version", VERSION_FILL, VERSION_FONT),
    ("Sprint - Tarea individual con duracion definida", VERSION_FILL, VERSION_FONT),
    ("Ciclo - Gran fase del proyecto (agrupacion de fases)", VERSION_FILL, VERSION_FONT),
]
for text, fill, font in legend_items:
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = font
    cell.fill = fill
    cell.border = thin_border
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    for c in range(2, 5):
        ws.cell(row=row, column=c).border = thin_border
        ws.cell(row=row, column=c).fill = fill
    row += 1

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:J{row - 10}"
for r in range(2, row):
    ws.row_dimensions[r].height = 24
ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.page_setup.orientation = "landscape"

output_path = f"{BASE}/Gantt_VERIFEX.xlsx"
wb.save(output_path)
print(f"Excel generado: {output_path}")

print("\nLinea de versiones del sistema:")
for td in task_defs:
    if len(td) >= 7 and td[0] == "V":
        name = td[2]
        dep = td[5]
        dep_key = f"T{dep}"
        if dep_key in task_dates:
            s, e = task_dates[dep_key]
            nd = e + timedelta(days=1)
            print(f"  {name:<59} -> {nd.strftime('%d/%m/%Y')}")
