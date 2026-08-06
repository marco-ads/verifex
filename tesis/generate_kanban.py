#!/usr/bin/env python3
"""Generador de Kanban VERIFEX - Horizontal"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = "/Users/maarco_serrano/Downloads/verifex-standalone 2"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Kanban VERIFEX"

# ── Colores por persona ──
MARCO_FILL = PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")
MARCO_FONT = Font(name="Calibri", size=9, bold=True, color="4A148C")
TONY_FILL = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
TONY_FONT = Font(name="Calibri", size=9, bold=True, color="B71C1C")
LUIS_FILL = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
LUIS_FONT = Font(name="Calibri", size=9, bold=True, color="0D47A1")
ULISES_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
ULISES_FONT = Font(name="Calibri", size=9, bold=True, color="1B5E20")

# ── Colores de columnas Kanban ──
COLORS = {
    "Backlog":       PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"),
    "Stories":       PatternFill(start_color="E8EAF6", end_color="E8EAF6", fill_type="solid"),
    "Por Hacer":     PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"),
    "En Proceso":    PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid"),
    "Por Verificar": PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"),
    "Acabado/Terminado": PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
}

HEADER_COLORS = {
    "Backlog":       PatternFill(start_color="9E9E9E", end_color="9E9E9E", fill_type="solid"),
    "Stories":       PatternFill(start_color="3F51B5", end_color="3F51B5", fill_type="solid"),
    "Por Hacer":     PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid"),
    "En Proceso":    PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid"),
    "Por Verificar": PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid"),
    "Acabado/Terminado": PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid"),
}

prioridad_colors = {
    "Alta":  Font(name="Calibri", size=9, bold=True, color="C62828"),
    "Media": Font(name="Calibri", size=9, bold=True, color="F57F17"),
    "Baja":  Font(name="Calibri", size=9, bold=True, color="2E7D32"),
}

assignee_styles = {
    "Marco":  (MARCO_FILL, MARCO_FONT),
    "Luis":   (LUIS_FILL, LUIS_FONT),
    "Ulises": (ULISES_FILL, ULISES_FONT),
    "Tony":   (TONY_FILL, TONY_FONT),
}

CHAPTER_FILL = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
CHAPTER_FONT = Font(name="Calibri", size=9, bold=True, color="1565C0")

VERSION_FILL = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
VERSION_FONT = Font(name="Calibri", size=9, bold=True, color="E65100")

# ── Versiones del sistema (del Gantt) ──
VERSIONS = [
    ("★ v0.1.0", "v0.1.0 — Requisitos, HU, marco teórico y plan de pruebas (04/12/2025)"),
    ("★ v0.2.0", "v0.2.0 — Diseño completo: arquitectura, prototipo, diagramas UML (12/01/2026)"),
    ("★ v0.3.0", "v0.3.0 — Scraper multi-estrategia funcional (09/02/2026)"),
    ("★ v0.4.0", "v0.4.0 — Clasificador de credibilidad e integración frontend-backend (07/03/2026)"),
    ("★ v0.5.0", "v0.5.0 — Desarrollo completo del sistema (31/03/2026)"),
    ("★ v0.6.0", "v0.6.0 — Suites completas de pruebas backend y frontend (19/04/2026)"),
    ("★ v0.7.0", "v0.7.0 — Configuración de despliegue y capturas del sistema (02/05/2026)"),
    ("★ v0.8.0", "v0.8.0 — Versión desplegada en Render con dominio público (15/05/2026)"),
    ("★ v0.9.0", "v0.9.0 — Validación post-despliegue y rendimiento verificado (21/05/2026)"),
    ("★ v0.10.0","v0.10.0 — Análisis con URLs reales y diseño responsive (03/06/2026)"),
    ("★ v1.0.0", "v1.0.0 — Documentación completa, manuales y versión estable (16/06/2026)"),
    ("★ v1.1.0", "v1.1.0 — Revisiones finales, verificación y maquetación (29/06/2026)"),
    ("★ v1.2.0", "v1.2.0 — Tesis completa, sistema final y defensa [PENDIENTE — 21/08/2026]"),
]

# A qué capítulo pertenece cada versión (para ordenarla en el Kanban)
version_chapter = {
    "v0.1.0": 1,  # Despues de Cap 2 (Fundamento Teorico — cierre de investigacion)
    "v0.2.0": 2,
    "v0.3.0": 2,
    "v0.4.0": 2,
    "v0.5.0": 2,
    "v0.6.0": 3,
    "v0.7.0": 3,
    "v0.8.0": 3,
    "v0.9.0": 3,
    "v0.10.0": 3,
    "v1.0.0": 3,
    "v1.1.0": 3,
    "v1.2.0": 4,
}

# ── Indice de la tesis: orden de capitulos ──
THESIS_CHAPTERS = [
    ("01", "Capítulo 1: Generalidades del Proyecto"),
    ("02", "Capítulo 2: Fundamento Teórico"),
    ("03", "Capítulo 3: Metodología y Desarrollo del Sistema"),
    ("04", "Capítulo 4: Implementación y Resultados"),
    ("05", "Conclusiones"),
]

# Mapeo de cada tarea a su capitulo de tesis
# task_id -> indice en THESIS_CHAPTERS
task_chapter = {
    # Cap 1: Generalidades
    "DOC-01": 0, "DOC-05": 0, "RF-01": 0, "TES-01": 0,
    # Cap 2: Fundamento Teorico
    "DOC-02": 1, "INV-01": 1, "TES-02": 1,
    "B-10": 1, "B-11": 1, "B-12": 1, "B-13": 1, "B-14": 1,
    # Cap 3: Metodologia de Desarrollo
    "DIS-01": 2, "DIS-02": 2, "DIS-03": 2, "DIS-04": 2,
    "DEV-01": 2, "DEV-02": 2, "DEV-03": 2, "DEV-04": 2,
    "DEV-05": 2, "DEV-06": 2, "DEV-07": 2, "DEV-08": 2,
    "DEV-09": 2, "DEV-10": 2, "DEV-11": 2, "DEV-12": 2,
    "DEV-13": 2, "DEV-14": 2, "DEV-15": 2,
    "TES-03": 2,
    # Cap 1: HU de requisitos (HU-01..02)
    "HU-01": 0, "HU-02": 0,
    # Cap 2: HU de investigacion (HU-03..07)
    "HU-03": 1, "HU-04": 1, "HU-05": 1, "HU-06": 1, "HU-07": 1,
    # Cap 3: HU de desarrollo (HU-08..29)
    "HU-08": 2, "HU-09": 2, "HU-10": 2,
    "HU-11": 2, "HU-12": 2, "HU-13": 2, "HU-14": 2,
    "HU-15": 2, "HU-16": 2, "HU-17": 2, "HU-18": 2,
    "HU-19": 2, "HU-20": 2, "HU-21": 2, "HU-22": 2,
    "HU-23": 2, "HU-24": 2, "HU-25": 2, "HU-26": 2,
    "HU-27": 2, "HU-28": 2, "HU-29": 2,
    # Cap 4: HU de pruebas, documentacion y despliegue (HU-30..39)
    "HU-30": 3, "HU-31": 3, "HU-32": 3, "HU-33": 3, "HU-34": 3,
    "HU-35": 3, "HU-36": 3, "HU-37": 3, "HU-38": 3, "HU-39": 3,
    # Cap 4: Implementacion y Resultados
    "TEST-01": 3, "TEST-02": 3, "TEST-03": 3,
    "DEP-01": 3, "DEP-02": 3, "DEP-03": 3, "DEP-04": 3,
    "RES-01": 3,
    "DIS-05": 3,
    "DOC-08": 3, "DOC-09": 3,
    "TES-04": 3,
    # Conclusiones
    "DOC-03": 4, "DOC-04": 4, "DOC-07": 4, "DOC-10": 4,
    "PDF-01": 4, "TESIS-01": 4, "DOC-06": 4,
    # Gestion del Proyecto (bajo Capitulo 4: Implementacion y Resultados)
    "GAN-00": 3, "GAN-01": 3, "KAN-01": 3,
    "SCR-01": 3, "SCR-02": 3,
    "COR-01": 3, "COR-02": 3, "COR-03": 3, "COR-04": 3,
    # Backlog: ideas futuras (bajo Capitulo 3: Desarrollo)
    "B-01": 2, "B-02": 2, "B-03": 2, "B-04": 2,
    "B-05": 2, "B-06": 2, "B-07": 2, "B-08": 2, "B-09": 2,
}

# Posicion de ordenamiento dentro de cada capitulo
# Permite intercalar versiones entre grupos logicos de tareas
item_position = {
    # Versiones (0 = al inicio del capitulo)
    "v0.1.0": 0,
    "v0.2.0": 0,
    "v0.3.0": 100,
    "v0.4.0": 200,
    "v0.5.0": 300,
    "v0.6.0": 0,
    "v0.7.0": 100,
    "v0.8.0": 200,
    "v0.9.0": 300,
    "v0.10.0": 400,
    "v1.0.0": 500,
    "v1.1.0": 600,
    "v1.2.0": 0,
    # Cap 1: Generalidades
    "HU-01": 5, "HU-02": 10,
    "DOC-01": 15, "RF-01": 20, "TES-01": 25,
    # Cap 2: Fundamento Teorico
    "HU-03": 5, "HU-04": 10, "HU-05": 12, "HU-06": 14, "HU-07": 16,
    "DOC-02": 30, "INV-01": 40, "TES-02": 50,
    "B-10": 100, "B-11": 110, "B-12": 120, "B-13": 130, "B-14": 140,
    # Cap 3 — HU secuenciales 01-39
    "HU-01": 1, "HU-02": 3,
    # Cap 3 — HU de desarrollo 08-29
    "HU-08": 5, "HU-09": 8, "HU-10": 11,
    "HU-11": 14, "HU-12": 17, "HU-13": 20, "HU-14": 23,
    "HU-15": 26, "HU-16": 29, "HU-17": 32, "HU-18": 35,
    "HU-19": 38, "HU-20": 41, "HU-21": 44, "HU-22": 47,
    "HU-23": 50, "HU-24": 53, "HU-25": 56, "HU-26": 59,
    "HU-27": 62, "HU-28": 65, "HU-29": 68,

    # Cap 3 — Diseno (entre v0.2.0 y v0.3.0)
    "DIS-01": 85, "DIS-02": 87, "DIS-03": 89, "DIS-04": 91, "DIS-05": 93,
    # Cap 3 — Scraper/Backend (entre v0.3.0 y v0.4.0)
    "DEV-01": 110, "DEV-02": 115, "DEV-03": 120, "DEV-04": 125,
    "DEV-05": 130, "DEV-06": 135, "DEV-07": 140, "DEV-08": 145,
    # Cap 3 — Frontend/IA (entre v0.4.0 y v0.5.0)
    "DEV-09": 210, "DEV-10": 215, "DEV-11": 220, "DEV-12": 225,
    "DEV-13": 230, "DEV-14": 235, "DEV-15": 240,
    # Cap 3 — Documentacion desarrollo (despues de v0.5.0)
    "DOC-04": 310, "DOC-08": 320, "DOC-09": 330, "TES-03": 340,
    # Cap 4 — HU de pruebas, docs y despliegue
    "HU-30": 1, "HU-31": 2, "HU-32": 3, "HU-33": 4, "HU-34": 5,
    "HU-35": 6, "HU-36": 7, "HU-37": 8, "HU-38": 9, "HU-39": 10,
    # Cap 4 — Pruebas (entre v0.6.0 y v0.7.0)
    "TEST-01": 20, "TEST-02": 25, "TEST-03": 30, "RES-01": 35, "DEP-04": 40,
    # Cap 4 — Configuracion despliegue (entre v0.7.0 y v0.8.0)
    "DEP-01": 110,
    # Cap 4 — Despliegue prod (entre v0.8.0 y v0.9.0)
    "DEP-02": 210, "DEP-03": 215,
    # Cap 4 — URLs reales / resultados (entre v0.9.0 y v1.0.0)
    "TES-04": 410,
    # Cap 4 — Gestion y scripts (entre v1.0.0 y v1.1.0)
    "GAN-00": 510, "GAN-01": 520, "KAN-01": 530,
    "SCR-01": 540, "SCR-02": 550,
    "COR-01": 560, "COR-02": 565, "COR-03": 570, "COR-04": 575,
    # Cap 5: Conclusiones
    "DOC-03": 10, "DOC-07": 20,
}

def sort_key(t):
    """Clave de ordenamiento: (capitulo, posicion)."""
    tid = t[0]
    if tid and isinstance(tid, str) and tid.startswith("★ v"):
        vname = tid.replace("★ ", "")
        return (version_chapter.get(vname, 99), item_position.get(vname, 999))
    if tid and isinstance(tid, str) and tid.startswith("★"):
        return (-1, 0)
    return (task_chapter.get(tid, 99), item_position.get(tid, 999))

def insert_chapter_headers(task_list):
    """Inserta headers de TODOS los capitulos de la tesis, mas las versiones
    y tareas correspondientes dentro de cada capitulo.
    Muestra la estructura completa aunque un capitulo no tenga tareas."""
    if not task_list:
        # Sin tareas, solo mostrar headers de capitulo
        items = []
    else:
        items = []
        for t in task_list:
            tid = t[0]
            if tid and isinstance(tid, str) and tid.startswith("★") and not tid.startswith("★ v"):
                continue  # descartar chapter headers (se regeneran)
            items.append(t)
        items.sort(key=sort_key)
    # Agrupar por capitulo
    by_ch = {}
    for t in items:
        tid = t[0]
        if tid and isinstance(tid, str) and tid.startswith("★ v"):
            vname = tid.replace("★ ", "")
            ch = version_chapter.get(vname, 99)
        else:
            ch = task_chapter.get(tid, 99)
        by_ch.setdefault(ch, []).append(t)
    result = []
    for ch_idx in range(len(THESIS_CHAPTERS)):
        result.append((f"★ {THESIS_CHAPTERS[ch_idx][0]}", THESIS_CHAPTERS[ch_idx][1], None))
        if ch_idx in by_ch:
            result.extend(by_ch[ch_idx])
    return result

col_labels = {
    "Backlog": "Backlog",
    "Por Hacer": "Por Hacer [WIP: 5]",
    "En Proceso": "En Proceso [WIP: 3]",
    "Por Verificar": "Por Verificar [WIP: 3]",
}

thin_border = Border(
    left=Side(style="thin", color="BDBDBD"),
    right=Side(style="thin", color="BDBDBD"),
    top=Side(style="thin", color="BDBDBD"),
    bottom=Side(style="thin", color="BDBDBD"),
)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

# ── Columnas Kanban ──
columnas = ["Backlog", "Stories", "Por Hacer", "En Proceso", "Por Verificar", "Acabado/Terminado"]

# ── Tareas por columna ──
# Cada tarea: (id, descripcion, asignado, prioridad) — en Acabado: (id, descripcion, asignado)
# Las tareas estan ordenadas por indice de la tesis
tasks = {
    "Backlog": insert_chapter_headers([
        ("DOC-05", "Redaccion de introduccion, resumen y abstract de la tesis", None, "Media"),
        ("DOC-06", "Maquetacion y formato del documento Word (estilos, indices, referencias, portada)", None, "Media"),
        ("B-01", "Autenticacion de usuarios (login/registro)", None, "Baja"),
        ("B-02", "Historial de analisis por usuario", None, "Baja"),
        ("B-03", "Plugin de navegador (extension Chrome/Firefox)", None, "Baja"),
        ("B-04", "Dashboard de estadisticas de desinformacion", None, "Baja"),
        ("B-05", "API REST publica para terceros", None, "Baja"),
        ("B-06", "Analisis de imagenes (deteccion de deepfakes)", None, "Baja"),
        ("B-07", "Exportacion de resultados (PDF, CSV)", None, "Baja"),
        ("B-08", "Integracion con redes sociales (bot de Twitter/Telegram)", None, "Baja"),
        ("B-09", "Notificaciones de noticias falsas en tiempo real", None, "Baja"),
        ("B-10", "Investigacion de nuevas fuentes de verificacion (fact-checkers adicionales)", None, "Baja"),
        ("B-11", "Estudio comparativo de APIs de fact-checking (Google Fact Check Tools, ClaimBuster)", None, "Baja"),
        ("B-12", "Ampliacion del marco teorico sobre desinformacion en Mexico y America Latina", None, "Baja"),
        ("B-13", "Investigacion de patrones linguisticos de desinformacion para mejorar clasificacion", None, "Baja"),
        ("B-14", "Estudio de tecnicas de NLP para deteccion de sesgos y clasificacion multimodal", None, "Baja"),
    ]),
    "Stories": insert_chapter_headers([
        # ── Analisis de Noticias ──
        ("HU-01", "[HU-01] Pegar URL de noticia para analizar su credibilidad", None, "Alta"),
        ("HU-02", "[HU-02] Ver veredicto claro (REAL/FALSO/ESTAFA/SATIRA/NO VERIFICABLE)", None, "Alta"),
        # ── Investigacion Teorica ──
        ("HU-35", "[HU-35] Investigar antecedentes de verificacion de noticias y fact-checkers existentes", None, "Media"),
        ("HU-36", "[HU-36] Revisar marco teorico sobre fake news, desinformacion y credibilidad digital", None, "Media"),
        ("HU-37", "[HU-37] Analizar viabilidad tecnica, economica y operativa del proyecto VERIFEX", None, "Media"),
        ("HU-38", "[HU-38] Estudiar herramientas del mercado: Google Fact Check, Snopes, Politifact, Verificado MX", None, "Media"),
        ("HU-39", "[HU-39] Definir metricas y criterios de clasificacion de credibilidad", None, "Media"),
        ("HU-03", "[HU-03] Ver nivel de confianza del analisis segmentado en 20 niveles", None, "Alta"),
        ("HU-04", "[HU-04] Ver banderas rojas y senales positivas detectadas en la noticia", None, "Media"),
        ("HU-05", "[HU-05] Ver resumen del articulo generado por IA con citas textuales", None, "Media"),
        ("HU-06", "[HU-06] Ver afirmaciones principales extraidas del contenido", None, "Media"),
        ("HU-07", "[HU-07] Ver analisis detallado por categoria con evidencias", None, "Media"),
        ("HU-08", "[HU-08] Ver noticias similares para contexto adicional y verificacion cruzada", None, "Media"),
        ("HU-09", "[HU-09] Ver tipo de articulo detectado (noticia, opinion, satira, editorial, presna amarillista)", None, "Media"),
        ("HU-10", "[HU-10] Recibir alerta automatica si la noticia es una estafa o fraude", None, "Alta"),

        # ── Idioma y Experiencia de Usuario ──
        ("HU-11", "[HU-11] Cambiar idioma entre espanol e ingles en toda la interfaz", None, "Media"),
        ("HU-12", "[HU-12] Interfaz facil de usar e intuitiva con diseno cyberpunk", None, "Media"),
        ("HU-13", "[HU-13] Ver indicador de carga animado mientras se analiza la URL", None, "Media"),
        ("HU-14", "[HU-14] Ver mensajes de error descriptivos segun el tipo de fallo (timeout, URL invalida, servidor caido)", None, "Media"),
        ("HU-15", "[HU-15] Diseno responsive adaptable a movil, tablet y escritorio", None, "Media"),

        # ── Privacidad y Seguridad ──
        ("HU-16", "[HU-16] El analisis no almacena ningun dato personal ni historial del usuario", None, "Alta"),
        ("HU-17", "[HU-17] Conexion segura via HTTPS/SSL con certificado valido", None, "Alta"),

        # ── Infraestructura Tecnica ──
        ("HU-18", "[HU-18] Extraccion robusta de contenido con scraper multi-estrategia (4 capas de fallback)", None, "Alta"),
        ("HU-19", "[HU-19] Clasificacion de credibilidad mediante IA (Groq API con fallback a llama-3.1-8b-instant)", None, "Alta"),
        ("HU-20", "[HU-20] Prompt engineering con 3 few-shots, citas textuales obligatorias y 5 categorias de veredicto", None, "Alta"),
        ("HU-21", "[HU-21] Lista de 29 dominios confiables con override automatico de clasificacion", None, "Media"),
        ("HU-22", "[HU-22] Busqueda de noticias similares via Google News RSS para contexto", None, "Media"),
        ("HU-23", "[HU-23] Sistema desplegado en Render con dominio publico y zero downtime", None, "Alta"),
        ("HU-24", "[HU-24] Backend en Flask con rutas /analyze y /health, CORS habilitado", None, "Alta"),
        ("HU-25", "[HU-25] Frontend en React 18 + TypeScript + Vite con conexion a backend via fetch + AbortController (60s timeout)", None, "Alta"),

        # ── Diseno Visual ──
        ("HU-26", "[HU-26] Diseno visual cyberpunk: cuadricula de fondo, glitch, vignette CRT, scanlines y clip-paths", None, "Media"),
        ("HU-27", "[HU-27] Componente ConfidenceBar con 20 niveles segmentados y codigo de colores", None, "Media"),
        ("HU-28", "[HU-28] Veredicto mostrado con indicador visual por color (verde=real, rojo=falso, ambar=sospechoso, azul=no verif)", None, "Media"),
        ("HU-29", "[HU-29] Componentes UI: UrlInput, VerdictDisplay, ConfidenceBar, RedFlags, SimilarNews, LanguageToggle", None, "Media"),

        # ── Documentacion ──
        ("HU-30", "[HU-30] Documentacion tecnica completa de la API, arquitectura y modulos del sistema", None, "Baja"),
        ("HU-31", "[HU-31] Manual de usuario con instrucciones de uso y ejemplos", None, "Baja"),
        ("HU-32", "[HU-32] Tesis documentada: introduccion, marco teorico, metodologia, implementacion y resultados", None, "Alta"),
        ("HU-33", "[HU-33] Diagramas UML completos: casos de uso, actividades, secuencia, clases", None, "Media"),
        ("HU-34", "[HU-34] Suites de pruebas automatizadas: pytest backend (27 tests) + vitest frontend (52 tests)", None, "Alta"),
    ]),
    "Por Hacer": insert_chapter_headers([
        ("TESIS-01", "Unificar criterios: documento Word y PDF deben coincidir en contenido", "Ulises", "Alta"),
    ]),
    "En Proceso": insert_chapter_headers([]),  # Tareas completadas segun Gantt
    "Por Verificar": insert_chapter_headers([
        ("DOC-10", "Correcciones finales y ajustes segun retroalimentacion", "Ulises", "Alta"),
        ("PDF-01", "Generacion de PDF final de tesis (TESIS_VERIFEX.pdf) con todas las correcciones", "Ulises", "Alta"),
    ]),
    "Acabado/Terminado": insert_chapter_headers([
        # Versiones del sistema (insertados como headers antes de cada capitulo)
        ("★ v0.1.0", "v0.1.0 — Requisitos, HU, marco teorico y plan de pruebas (04/12/2025)", None),
        ("★ v0.2.0", "v0.2.0 — Diseno completo: arquitectura, prototipo, diagramas UML (12/01/2026)", None),
        ("★ v0.3.0", "v0.3.0 — Scraper multi-estrategia funcional (09/02/2026)", None),
        ("★ v0.4.0", "v0.4.0 — Clasificador de credibilidad e integracion frontend-backend (07/03/2026)", None),
        ("★ v0.5.0", "v0.5.0 — Desarrollo completo del sistema (31/03/2026)", None),
        ("★ v0.6.0", "v0.6.0 — Suites completas de pruebas backend y frontend (19/04/2026)", None),
        ("★ v0.7.0", "v0.7.0 — Configuracion de despliegue y capturas del sistema (02/05/2026)", None),
        ("★ v0.8.0", "v0.8.0 — Version desplegada en Render con dominio publico (15/05/2026)", None),
        ("★ v0.9.0", "v0.9.0 — Validacion post-despliegue y rendimiento verificado (21/05/2026)", None),
        ("★ v0.10.0","v0.10.0 — Analisis con URLs reales y diseno responsive (03/06/2026)", None),
        ("★ v1.0.0", "v1.0.0 — Documentacion completa, manuales y version estable (16/06/2026)", None),
        ("★ v1.1.0", "v1.1.0 — Revisiones finales, verificacion y maquetacion (29/06/2026)", None),
        ("★ v1.2.0", "v1.2.0 — Tesis completa, sistema final y defensa [PENDIENTE — 21/08/2026]", None),

        # Cap 1: Generalidades del Proyecto
        ("DOC-01", "Eleccion del tema de tesis: deteccion de noticias falsas con IA", "Todos"),
        ("RF-01", "Definicion de requisitos funcionales (10 HU) y no funcionales (rendimiento, seguridad, usabilidad)", "Marco"),
        ("TES-01", "Capitulo 1: Introduccion (contexto, problema, objetivos, justificacion)", "Ulises"),

        # Cap 2: Fundamento Teorico
        ("DOC-02", "Investigacion preliminar y revision bibliografica (fake news, verificacion, IA)", "Todos"),
        ("INV-01", "Investigacion sobre IA, Groq API y procesamiento de lenguaje natural", "Marco"),
        ("TES-02", "Capitulo 2: Marco Teorico (fake news, IA, Groq API, verificacion)", "Ulises"),
        ("HU-35", "[HU-35] Investigar antecedentes de verificacion de noticias y fact-checkers existentes", "Todos"),
        ("HU-36", "[HU-36] Revisar marco teorico sobre fake news, desinformacion y credibilidad digital", "Todos"),
        ("HU-37", "[HU-37] Analizar viabilidad tecnica, economica y operativa del proyecto VERIFEX", "Marco"),
        ("HU-38", "[HU-38] Estudiar herramientas del mercado: Google Fact Check, Snopes, Politifact, Verificado MX", "Todos"),
        ("HU-39", "[HU-39] Definir metricas y criterios de clasificacion de credibilidad", "Marco"),

        # Cap 3: Metodologia de Desarrollo
        ("HU-01", "[HU-01] Pegar URL de noticia para analizar su credibilidad", "Marco"),
        ("HU-02", "[HU-02] Ver veredicto claro (REAL/FALSO/ESTAFA/SATIRA/NO VERIFICABLE)", "Marco"),
        ("HU-03", "[HU-03] Ver nivel de confianza del analisis segmentado en 20 niveles", "Marco"),
        ("HU-04", "[HU-04] Ver banderas rojas y senales positivas detectadas en la noticia", "Marco"),
        ("HU-05", "[HU-05] Ver resumen del articulo generado por IA con citas textuales", "Marco"),
        ("HU-06", "[HU-06] Ver afirmaciones principales extraidas del contenido", "Marco"),
        ("HU-07", "[HU-07] Ver analisis detallado por categoria con evidencias", "Marco"),
        ("HU-08", "[HU-08] Ver noticias similares para contexto adicional y verificacion cruzada", "Marco"),
        ("HU-09", "[HU-09] Ver tipo de articulo detectado (noticia, opinion, satira, editorial, prensa amarillista)", "Marco"),
        ("HU-10", "[HU-10] Recibir alerta automatica si la noticia es una estafa o fraude", "Marco"),
        ("HU-11", "[HU-11] Cambiar idioma entre espanol e ingles en toda la interfaz", "Luis"),
        ("HU-12", "[HU-12] Interfaz facil de usar e intuitiva con diseno cyberpunk", "Luis"),
        ("HU-13", "[HU-13] Ver indicador de carga animado mientras se analiza la URL", "Luis"),
        ("HU-14", "[HU-14] Ver mensajes de error descriptivos segun el tipo de fallo (timeout, URL invalida, servidor caido)", "Luis"),
        ("HU-15", "[HU-15] Diseno responsive adaptable a movil, tablet y escritorio", "Luis"),
        ("HU-16", "[HU-16] El analisis no almacena ningun dato personal ni historial del usuario", "Marco"),
        ("HU-17", "[HU-17] Conexion segura via HTTPS/SSL con certificado valido", "Marco"),
        ("HU-18", "[HU-18] Extraccion robusta de contenido con scraper multi-estrategia (4 capas de fallback)", "Marco"),
        ("HU-19", "[HU-19] Clasificacion de credibilidad mediante IA (Groq API con fallback a llama-3.1-8b-instant)", "Marco"),
        ("HU-20", "[HU-20] Prompt engineering con 3 few-shots, citas textuales obligatorias y 5 categorias de veredicto", "Marco"),
        ("HU-21", "[HU-21] Lista de 29 dominios confiables con override automatico de clasificacion", "Marco"),
        ("HU-22", "[HU-22] Busqueda de noticias similares via Google News RSS para contexto", "Marco"),
        ("HU-23", "[HU-23] Sistema desplegado en Render con dominio publico y zero downtime", "Marco"),
        ("HU-24", "[HU-24] Backend en Flask con rutas /analyze y /health, CORS habilitado", "Marco"),
        ("HU-25", "[HU-25] Frontend en React 18 + TypeScript + Vite con conexion a backend via fetch + AbortController (60s timeout)", "Marco"),
        ("HU-26", "[HU-26] Diseno visual cyberpunk: cuadricula de fondo, glitch, vignette CRT, scanlines y clip-paths", "Luis"),
        ("HU-27", "[HU-27] Componente ConfidenceBar con 20 niveles segmentados y codigo de colores", "Luis"),
        ("HU-28", "[HU-28] Veredicto mostrado con indicador visual por color (verde=real, rojo=falso, ambar=sospechoso, azul=no verif)", "Luis"),
        ("HU-29", "[HU-29] Componentes UI: UrlInput, VerdictDisplay, ConfidenceBar, RedFlags, SimilarNews, LanguageToggle", "Luis"),
        ("HU-30", "[HU-30] Documentacion tecnica completa de la API, arquitectura y modulos del sistema", "Tony"),
        ("HU-31", "[HU-31] Manual de usuario con instrucciones de uso y ejemplos", "Luis"),
        ("HU-32", "[HU-32] Tesis documentada: introduccion, marco teorico, metodologia, implementacion y resultados", "Ulises"),
        ("HU-33", "[HU-33] Diagramas UML completos: casos de uso, actividades, secuencia, clases", "Tony"),
        ("HU-34", "[HU-34] Suites de pruebas automatizadas: pytest backend (27 tests) + vitest frontend (52 tests)", "Marco"),
        ("DIS-01", "Wireframes y mockups de interfaz de usuario", "Luis"),
        ("DIS-02", "Paleta de colores y fuentes cyberpunk (Orbitron, Rajdhani, Share Tech Mono)", "Luis"),
        ("DIS-03", "Arquitectura cliente-servidor (React + Flask + Groq API)", "Marco"),
        ("DIS-04", "Diagramas UML (casos de uso, actividades, secuencia, clases, entidad-relacion)", "Tony"),
        ("DIS-05", "Capturas de pantalla del sistema (interfaz, analisis, resultados, errores, mockups)", "Luis"),
        ("DEV-01", "Setup del proyecto: Vite + React 18 + TypeScript + Tailwind CSS + PostCSS", "Marco"),
        ("DEV-02", "Setup del backend: Flask + CORS + rutas /analyze y /health", "Marco"),
        ("DEV-03", "Scraper URLs con fallback 4-capas: cloudscraper -> curl_cffi -> requests -> Playwright (Firefox)", "Marco"),
        ("DEV-04", "Integracion con Groq API (call_groq con fallback a llama-3.1-8b-instant)", "Marco"),
        ("DEV-05", "Prompt engineering: citas textuales obligatorias, 3 few-shot, banderas rojas especificas, 5 categorias", "Marco"),
        ("DEV-06", "Clasificador de credibilidad (analyze_url: REAL, FALSO, SATIRA, ESTAFA, NO VERIFICABLE)", "Marco"),
        ("DEV-07", "Lista de dominios confiables (CREDIBLE_DOMAINS con 29 fuentes noticiosas)", "Marco"),
        ("DEV-08", "Verificacion cruzada via Google News RSS (news_finder.py + find_similar_news)", "Marco"),
        ("DEV-09", "Conexion frontend-backend (fetch /analyze con AbortController 60s)", "Marco"),
        ("DEV-10", "Componentes base: UrlInput, VerdictDisplay, ConfidenceBar, RedFlags", "Luis"),
        ("DEV-11", "Componentes: SimilarNews, LanguageToggle, estados (loading/error/results)", "Luis"),
        ("DEV-12", "Soporte bilingue ES/EN en todos los componentes", "Luis"),
        ("DEV-13", "Article_type (5 tipos) y deteccion de estafas (is_scam)", "Marco"),
        ("DEV-14", "Diseno visual cyberpunk: cuadricula, glitch, vignette CRT, scanlines, clip-paths", "Luis"),
        ("DEV-15", "Creacion de build.sh para instalar Firefox durante el build de Render", "Marco"),
        ("DOC-04", "Correccion del documento Word (acentos, enie, formato, tabla de contenido)", "Ulises"),
        ("DOC-08", "Redaccion del Capitulo 5: Resultados, pruebas y conclusiones", "Ulises"),
        ("DOC-09", "Analisis de casos de prueba y documentacion de resultados en la tesis", "Tony"),
        ("TES-03", "Capitulo 3: Metodologia y Diseno del Sistema", "Ulises"),

        # Cap 4: Implementacion y Resultados
        ("RES-01", "Pruebas de analisis con URLs reales (positivas, negativas, estafas, satira, opinion)", "Marco"),
        ("DEP-04", "Correccion de errores: timeouts (AbortController 60s), override de dominios confiables, edge cases", "Marco"),
        ("TEST-01", "Suite de pruebas backend (pytest, 27 tests en server/test_analyzer.py)", "Marco"),
        ("TEST-02", "Suite de pruebas frontend (vitest + testing-library, 7 archivos, 52 tests)", "Marco"),
        ("TEST-03", "Pruebas de integracion: frontend + backend + Groq API en conjunto", "Marco"),
        ("DEP-01", "Configuracion de despliegue: Procfile con gunicorn, variables de entorno, CORS", "Marco"),
        ("DEP-02", "Despliegue en Railway (fallo por limitaciones de memoria), migracion a Render", "Marco"),
        ("DEP-03", "Despliegue exitoso en Render con dominio publico y certificado SSL", "Marco"),
        ("TES-04", "Capitulo 4: Desarrollo e Implementacion", "Ulises"),

        # Gestion del Proyecto
        ("GAN-00", "Creacion del diagrama de Gantt del proyecto (Gantt_VERIFEX.xlsx, 49 tareas, 12 hitos)", "Ulises"),
        ("GAN-01", "Actualizacion del Gantt con fechas (fin 15/08/2026) y reubicacion de Gantt/Kanban", "Ulises"),
        ("KAN-01", "Creacion del tablero Kanban del proyecto (Kanban_VERIFEX.xlsx)", "Todos"),
        ("SCR-01", "Script generate_gantt_excel.py para automatizar el diagrama de Gantt", "Ulises"),
        ("SCR-02", "Script generate_thesis.py para automatizar la generacion de la tesis", "Ulises"),
        ("COR-01", "Correccion de tesis: Ollama -> Groq API, local -> nube, Detector Fake News -> Verifex", "Ulises"),
        ("COR-02", "Correccion de acentos (stripped) y conservacion de enie en toda la tesis", "Ulises"),
        ("COR-03", "Correccion de referencias a 'phi3:mini', 'mistral', '4GB modelo local' en la tesis", "Ulises"),
        ("COR-04", "Correccion de prompts con cita textual, few-shot y distincion opinion/informativa", "Ulises"),

        # Conclusiones
        ("DOC-03", "Verificar que no queden referencias a 'Ollama', 'IA local' o 'sin conexion' en el Word", "Ulises"),
        ("DOC-07", "Revision de contenido, ortografia y consistencia de la tesis", "Tony"),
    ]),
}

# ── Construir hoja horizontal ──
# Fila 1: titulo del proyecto
ws.merge_cells("A1:V1")
title = ws.cell(row=1, column=1, value="KANBAN - VERIFEX: Analizador de Credibilidad de Noticias")
title.font = Font(name="Calibri", bold=True, size=16, color="1A237E")
title.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 35

# Fila 2: encabezados de columnas Kanban
col_inicio = {}
col_idx = 1
for col_name in columnas:
    col_inicio[col_name] = col_idx
    ncols = 3  # ID, Tarea, Asignado/Prioridad
    ws.merge_cells(start_row=2, start_column=col_idx, end_row=2, end_column=col_idx + ncols - 1)
    cell = ws.cell(row=2, column=col_idx, value=col_labels.get(col_name, col_name))
    cell.font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    cell.fill = HEADER_COLORS[col_name]
    cell.alignment = center
    cell.border = thin_border
    for c in range(col_idx, col_idx + ncols):
        ws.cell(row=2, column=c).border = thin_border
        ws.cell(row=2, column=c).fill = HEADER_COLORS[col_name]
    # Subheaders
    ws.cell(row=3, column=col_idx, value="ID").font = Font(name="Calibri", bold=True, size=9, color="FFFFFF")
    ws.cell(row=3, column=col_idx).fill = HEADER_COLORS[col_name]
    ws.cell(row=3, column=col_idx).alignment = center
    ws.cell(row=3, column=col_idx).border = thin_border
    ws.cell(row=3, column=col_idx + 1, value="Tarea / Descripcion").font = Font(name="Calibri", bold=True, size=9, color="FFFFFF")
    ws.cell(row=3, column=col_idx + 1).fill = HEADER_COLORS[col_name]
    ws.cell(row=3, column=col_idx + 1).alignment = center
    ws.cell(row=3, column=col_idx + 1).border = thin_border
    ws.cell(row=3, column=col_idx + 2, value="Asignado / Prio").font = Font(name="Calibri", bold=True, size=9, color="FFFFFF")
    ws.cell(row=3, column=col_idx + 2).fill = HEADER_COLORS[col_name]
    ws.cell(row=3, column=col_idx + 2).alignment = center
    ws.cell(row=3, column=col_idx + 2).border = thin_border
    col_idx += ncols

# Anchos de columna
col_idx = 1
for col_name in columnas:
    ws.column_dimensions[get_column_letter(col_idx)].width = 8
    ws.column_dimensions[get_column_letter(col_idx + 1)].width = 52
    ws.column_dimensions[get_column_letter(col_idx + 2)].width = 16
    col_idx += 3

# Rellenar tareas
max_filas = max(len(tasks[c]) for c in columnas)
for fila_offset in range(max_filas):
    fila = fila_offset + 4
    col_idx = 1
    for col_name in columnas:
        lista = tasks[col_name]
        if fila_offset < len(lista):
            t = lista[fila_offset]
            tid, desc, asignado, *resto = t
            prioridad = resto[0] if resto else None

            # ID
            c_id = ws.cell(row=fila, column=col_idx, value=tid)
            c_id.font = Font(name="Calibri", size=9, bold=True)
            c_id.alignment = center
            c_id.border = thin_border
            c_id.fill = COLORS[col_name]

            # Descripcion
            c_desc = ws.cell(row=fila, column=col_idx + 1, value=desc)
            c_desc.font = Font(name="Calibri", size=9)
            c_desc.alignment = left_wrap
            c_desc.border = thin_border
            c_desc.fill = COLORS[col_name]

            # Asignado / Prioridad
            c_asig = ws.cell(row=fila, column=col_idx + 2)
            c_asig.border = thin_border
            c_asig.fill = COLORS[col_name]
            c_asig.alignment = center

            # Header: version (★ v) or chapter (★ NN) - applies to all columns
            if tid and isinstance(tid, str) and tid.startswith("★"):
                is_version = len(tid) > 2 and tid[2:3] == 'v'
                ws.merge_cells(start_row=fila, start_column=col_idx, end_row=fila, end_column=col_idx + 2)
                cell = ws.cell(row=fila, column=col_idx, value=f"{tid}  {desc}")
                cell.font = VERSION_FONT if is_version else CHAPTER_FONT
                cell.fill = VERSION_FILL if is_version else CHAPTER_FILL
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = thin_border
                for cc in range(col_idx, col_idx + 3):
                    ws.cell(row=fila, column=cc).border = thin_border
                    ws.cell(row=fila, column=cc).fill = VERSION_FILL if is_version else CHAPTER_FILL
            elif col_name == "Acabado/Terminado":
                c_asig.value = asignado if asignado else "—"
                if asignado and asignado in assignee_styles:
                    a_fill, a_font = assignee_styles[asignado]
                    c_asig.fill = a_fill
                    c_asig.font = a_font
                else:
                    c_asig.font = Font(name="Calibri", size=9)
            else:
                if asignado and prioridad and prioridad != "Baja":
                    c_asig.value = f"{asignado} [{prioridad}]"
                elif asignado:
                    c_asig.value = asignado
                elif prioridad:
                    c_asig.value = f"[{prioridad}]"
                else:
                    c_asig.value = "—"

                if asignado and asignado in assignee_styles:
                    a_fill, a_font = assignee_styles[asignado]
                    c_asig.fill = a_fill
                    c_asig.font = a_font
                elif prioridad and prioridad in prioridad_colors:
                    c_asig.font = prioridad_colors[prioridad]

        else:
            for c in range(3):
                cell = ws.cell(row=fila, column=col_idx + c)
                cell.border = thin_border
                cell.fill = COLORS[col_name]

        col_idx += 3

# Altura de filas
for r in range(2, max_filas + 4):
    ws.row_dimensions[r].height = 28 if r <= 3 else 36

# ── Leyenda al final ──
fila_leyenda = max_filas + 7
ws.merge_cells(start_row=fila_leyenda, start_column=1, end_row=fila_leyenda, end_column=6)
ws.cell(row=fila_leyenda, column=1, value="Leyenda:").font = Font(name="Calibri", bold=True, size=11)
fila_leyenda += 1

legend_items = [
    ("Marco - Backend, frontend, IA, deploy y pruebas", MARCO_FILL, MARCO_FONT),
    ("Luis - Frontend, diseno UI/UX y estilos visuales", LUIS_FILL, LUIS_FONT),
    ("Ulises - Documentacion y redaccion de tesis", ULISES_FILL, ULISES_FONT),
    ("Tony - Diagramas UML, documentacion y revision de logica", TONY_FILL, TONY_FONT),
    ("★ vX.Y.Z - Version o hito de lanzamiento", VERSION_FILL, VERSION_FONT),
    ("★ Capitulo N - Indice de la tesis", CHAPTER_FILL, CHAPTER_FONT),
    ("[Alta] Prioridad alta", None, prioridad_colors["Alta"]),
    ("[Media] Prioridad media", None, prioridad_colors["Media"]),
    ("[Baja] Prioridad baja", None, prioridad_colors["Baja"]),
]
for text, fill, font in legend_items:
    cell = ws.cell(row=fila_leyenda, column=1, value=text)
    cell.font = font
    if fill:
        cell.fill = fill
    cell.border = thin_border
    ws.merge_cells(start_row=fila_leyenda, start_column=1, end_row=fila_leyenda, end_column=3)
    for c in range(2, 4):
        ws.cell(row=fila_leyenda, column=c).border = thin_border
        if fill:
            ws.cell(row=fila_leyenda, column=c).fill = fill
    fila_leyenda += 1

ws.freeze_panes = "A4"
ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.page_setup.orientation = "landscape"

output_path = f"{BASE}/Kanban_VERIFEX.xlsx"
wb.save(output_path)
print(f"Kanban horizontal generado: {output_path}")
