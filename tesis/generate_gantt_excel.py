#!/usr/bin/env python3
"""
Generador de Gantt VERIFEX
- Tareas paralelas para Marco, Luis y Ulises en cada fase
- Todos los integrantes tienen actividad continua sin huecos
- Secuencial desde 01/09/2025, entrega 23/08/2026
- Dependencias lógicas: no se avanza de fase hasta que todos completan
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, timedelta

BASE = "/Users/maarco_serrano/Downloads/verifex-standalone 2"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Gantt VERIFEX"

MARCO_FILL = PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")
MARCO_FONT = Font(name="Calibri", size=10, bold=True, color="4A148C")
LUIS_FILL = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
LUIS_FONT = Font(name="Calibri", size=10, bold=True, color="0D47A1")
ULISES_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
ULISES_FONT = Font(name="Calibri", size=10, bold=True, color="1B5E20")

VERSION_FILL = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
VERSION_FONT = Font(name="Calibri", size=10, bold=True, color="E65100")

header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
phase_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
phase_fill = PatternFill(start_color="3949AB", end_color="3949AB", fill_type="solid")
thin_border = Border(
    left=Side(style="thin", color="BDBDBD"),
    right=Side(style="thin", color="BDBDBD"),
    top=Side(style="thin", color="BDBDBD"),
    bottom=Side(style="thin", color="BDBDBD"),
)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

headers = ["Fase", "Tipo", "#", "Tarea", "Asignado a", "Inicio", "Fin", "Dias", "Depende de", "Estado"]
col_widths = [32, 10, 5, 64, 14, 13, 13, 7, 14, 15]

for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(col_idx)].width = w

# Each entry: (marker, task_num, task_name, assignee, duration_days, depends_on, tipo)
# marker = None (regular task), "V" (version milestone)
# tipo = "Ciclo" (phase header), "Sprint" (regular task), "Iteracion" (milestone)
#
# Structure: each phase has parallel streams (one per person).
# Within each phase, everyone works simultaneously on their tasks.
# Dependencies: each person's task depends on their previous task.
# Phase transitions: all must finish before next phase starts.

task_defs = [
    # ══════════ CICLO 0: GESTION DE TESIS ══════════
    (None, None, None, None, None, None, None),

    (None, 1, "Eleccion del tema de tesis: deteccion de noticias falsas con IA", "Todos", 9, None, "Sprint"),
    (None, 2, "Investigacion preliminar y revision bibliografica (fake news, verificacion, IA)", "Todos", 22, 1, "Sprint"),
    (None, 3, "Definicion del problema, preguntas de investigacion y objetivos", "Todos", 5, 2, "Sprint"),
    (None, 4, "Aprobacion formal del tema de tesis y delimitacion del alcance", "Todos", 3, 3, "Sprint"),

    # ══════════ CICLO 1: FUNDAMENTOS TEORICOS ══════════
    (None, None, None, None, None, None, None),

    # Ronda 1: Investigacion paralela
    (None, 5, "Backend e IA: Groq API, Flask, scraping, curl_cffi, Playwright", "Marco", 18, 4, "Sprint"),
    (None, 6, "UI/UX cyberpunk: React, Tailwind, tipografia, paletas", "Luis", 17, 4, "Sprint"),
    (None, 7, "Marco teorico: fake news, verificacion digital, desinformacion MX", "Ulises", 19, 4, "Sprint"),
    (None, 8, "Herramientas de prueba: Vitest, Pytest, Playwright, Selenium", "Marco", 18, 4, "Sprint"),

    # Ronda 2: Planificacion paralela
    (None, 9, "Seleccion de tecnologias y definicion de requisitos funcionales", "Marco", 12, 5, "Sprint"),
    (None, 10, "Historias de usuario (HU-01 a HU-10) y diseno conceptual de interfaz", "Luis", 11, 6, "Sprint"),
    (None, 11, "Planteamiento del problema, justificacion y preguntas de investigacion", "Ulises", 13, 7, "Sprint"),
    (None, 12, "Metricas de calidad, criterios de aceptacion y matriz de trazabilidad", "Ulises", 12, 8, "Sprint"),

    # Ronda 3: Sintesis paralela
    (None, 13, "Setup de entorno: Node, Python, Git, VSCode, dependencias", "Marco", 10, 9, "Sprint"),
    (None, 14, "Moodboard, exploracion de componentes UI y paleta de colores", "Luis", 9, 10, "Sprint"),
    (None, 15, "Capitulo 1: Introduccion (contexto, problema, objetivos, alcance)", "Ulises", 11, 11, "Sprint"),
    (None, 16, "Plan maestro de pruebas y diseno de casos de prueba iniciales", "Ulises", 10, 12, "Sprint"),

    # Revision conjunta
    (None, 17, "Revision consolidacion: requisitos, HU y plan de pruebas", "Todos", 7, 16, "Sprint"),
    ("V", None, "v0.1.0 - Requisitos, HU, marco teorico y plan de pruebas definidos", None, 1, 17, "Iteracion"),

    # Revision post-Ciclo 1
    (None, 86, "Revision y retroalimentacion del Ciclo 1: ajustes a requisitos, HU y marco teorico", "Todos", 5, 17, "Sprint"),

    # ══════════ CICLO 2: DISENO DEL SISTEMA ══════════
    (None, None, None, None, None, None, None),

    # Ronda 1: Diseno inicial paralelo
    (None, 18, "Arquitectura cliente-servidor: React + Flask + Groq API + Playwright + CORS", "Marco", 12, 86, "Sprint"),
    (None, 19, "Wireframes baja fidelidad y prototipo navegable alta fidelidad", "Luis", 11, 86, "Sprint"),
    (None, 20, "Capitulo 2: Marco Teorico (fake news, IA, Groq API, verificacion)", "Ulises", 13, 86, "Sprint"),
    (None, 21, "Diagramas UML: casos de uso, actividades, secuencia, clases, ER", "Luis", 12, 86, "Sprint"),

    # Ronda 2: Setup y guias paralelo
    (None, 22, "Setup proyecto: Vite + React + TypeScript + Tailwind + PostCSS", "Marco", 8, 18, "Sprint"),
    (None, 23, "Paleta colores definitiva, tipografia (Orbitron, Rajdhani) y guia de estilos", "Luis", 7, 19, "Sprint"),
    (None, 24, "Diagrama Gantt y tablero Kanban (Gantt_VERIFEX.xlsx, Kanban_VERIFEX.xlsx)", "Ulises", 9, 20, "Sprint"),
    (None, 25, "Datos de prueba y configuracion del entorno de validacion", "Marco", 8, 21, "Sprint"),

    # Ronda 3: Diseno detallado paralelo
    (None, 26, "Diseno detallado del scraper multi-estrategia y plan de extraccion", "Marco", 10, 22, "Sprint"),
    (None, 27, "Maquetacion: UrlInput, layout inicial y estructura de paneles", "Luis", 9, 23, "Sprint"),
    (None, 28, "Capitulo 3: Metodologia y Diseno del Sistema (arquitectura, metodos)", "Ulises", 11, 24, "Sprint"),
    (None, 29, "Casos de prueba detallados para scraper y extraccion de contenido", "Marco", 10, 25, "Sprint"),

    ("V", None, "v0.2.0 - Diseno completo: arquitectura, prototipo navegable, diagramas UML y plan de pruebas detallado", None, 1, 29, "Iteracion"),

    # Revision post-Ciclo 2
    (None, 87, "Revision y retroalimentacion del Ciclo 2: validacion de diseno, prototipo y plan de pruebas", "Todos", 5, 29, "Sprint"),

    # ══════════ CICLO 3: DESARROLLO DEL SISTEMA ══════════
    (None, None, None, None, None, None, None),

    # Ronda 1: Core inicial paralelo
    (None, 30, "Setup Flask: rutas /analyze y /health, CORS, parse_response, errores", "Marco", 8, 87, "Sprint"),
    (None, 31, "Componentes base: VerdictDisplay, ConfidenceBar, veredicto por color", "Luis", 7, 87, "Sprint"),
    (None, 32, "Resultados esperados, metricas de evaluacion y criterios de exito", "Ulises", 9, 87, "Sprint"),
    (None, 33, "Configuracion de pruebas: pytest, vitest, jsdom, testing-library", "Marco", 8, 87, "Sprint"),

    # Ronda 2: Scraping + UI paralelo
    (None, 34, "Scraper multi-estrategia: cloudscraper, curl_cffi, requests, Playwright", "Marco", 12, 30, "Sprint"),
    (None, 35, "Componentes: RedFlags, SimilarNews, LanguageToggle con estilos", "Luis", 11, 31, "Sprint"),
    (None, 36, "Documentacion de avances del Capitulo 4: Desarrollo e Implementacion", "Ulises", 13, 32, "Sprint"),
    (None, 37, "Pruebas del scraper y validacion de extraccion de contenido HTML", "Marco", 12, 33, "Sprint"),

    ("V", None, "v0.3.0 - Scraper multi-estrategia funcional: cloudscraper, curl_cffi, requests y Playwright integrados y validados", None, 1, 37, "Iteracion"),

    # Ronda 3: IA + integracion paralelo
    (None, 38, "Llamadas a Groq API: call_groq, prompt engineering, parse_response", "Marco", 12, 34, "Sprint"),
    (None, 39, "Manejo de estados (loading, error, results) y soporte bilingue ES/EN", "Luis", 11, 35, "Sprint"),
    (None, 40, "Actualizacion Capitulos 1-3 segun cambios de implementacion", "Ulises", 13, 36, "Sprint"),
    (None, 41, "Pruebas de integracion scraper + Groq API con distintos tipos de URL", "Luis", 12, 37, "Sprint"),

    # Ronda 4: Clasificador + diseno final paralelo
    (None, 42, "Clasificador de credibilidad: 5 categorias, CREDIBLE_DOMAINS y override", "Marco", 12, 38, "Sprint"),
    (None, 43, "Diseno visual cyberpunk: cuadricula, glitch, vignette CRT, scanlines", "Luis", 11, 39, "Sprint"),
    (None, 44, "Tabla de resultados, analisis preliminar de datos e indice de tesis", "Ulises", 13, 40, "Sprint"),
    (None, 45, "Pruebas de integracion frontend-backend con datos simulados y reales", "Luis", 12, 41, "Sprint"),

    ("V", None, "v0.4.0 - Clasificador de credibilidad, integracion frontend-backend y diseno cyberpunk completos", None, 1, 45, "Iteracion"),

    # Ronda 5: Features avanzados paralelo
    (None, 46, "Busqueda de noticias similares: news_finder.py, Google News RSS", "Marco", 12, 42, "Sprint"),
    (None, 47, "Conexion frontend-backend via API REST (fetch /analyze + AbortController)", "Luis", 11, 43, "Sprint"),
    (None, 48, "Referencias bibliograficas, anexos y plantilla oficial de tesis", "Ulises", 13, 44, "Sprint"),
    (None, 49, "Pruebas de regresion y deteccion de defectos en integracion continua", "Luis", 12, 45, "Sprint"),

    # Ronda 6: Pulido final paralelo
    (None, 50, "Article_type: 5 tipos y ajustes finales de clasificacion", "Marco", 10, 46, "Sprint"),
    (None, 51, "Manejo de errores de red, timeouts y UX de carga animada", "Luis", 9, 47, "Sprint"),
    (None, 52, "Resultados parciales, graficas y tablas de visualizacion", "Ulises", 11, 48, "Sprint"),
    (None, 53, "Pruebas de aceptacion de usuario (UAT) con escenarios reales", "Ulises", 10, 49, "Sprint"),

    ("V", None, "v0.5.0 - Desarrollo completo: scraper, IA, clasificador, news_finder, UI cyberpunk y pruebas de integracion", None, 1, 53, "Iteracion"),

    # Revision post-Ciclo 3
    (None, 88, "Revision Ciclo 3: validacion del desarrollo completo y correcciones", "Todos", 5, 53, "Sprint"),

    # ══════════ CICLO 4: VALIDACION Y DESPLIEGUE ══════════
    (None, None, None, None, None, None, None),

    # Ronda 1: Suites de prueba paralelo
    (None, 54, "Suite pruebas backend: 27 tests en server/test_analyzer.py (pytest)", "Marco", 12, 88, "Sprint"),
    (None, 55, "Suite pruebas frontend: 52 tests con vitest + testing-library", "Luis", 11, 88, "Sprint"),
    (None, 56, "Resultados detallados y analisis de datos recopilados", "Ulises", 13, 88, "Sprint"),
    (None, 57, "Documentacion de pruebas: defectos, cobertura y metricas", "Ulises", 12, 88, "Sprint"),

    ("V", None, "v0.6.0 - Suites completas de pruebas backend y frontend, resultados documentados y metricas de cobertura", None, 1, 57, "Iteracion"),

    # Ronda 2: Despliegue paralelo
    (None, 58, "Configuracion despliegue: Procfile, build.sh, CORS, gunicorn, Railway", "Marco", 12, 54, "Sprint"),
    (None, 59, "Capturas del sistema: interfaz, resultados, errores, responsive", "Luis", 11, 55, "Sprint"),
    (None, 60, "Actualizacion bibliografia, graficas y tablas de visualizacion", "Ulises", 13, 56, "Sprint"),
    (None, 61, "Validacion de seguridad: manejo errores, proteccion API key, estres", "Marco", 12, 57, "Sprint"),

    ("V", None, "v0.7.0 - Configuracion de despliegue, capturas del sistema, validacion de seguridad y bibliografia actualizada", None, 1, 61, "Iteracion"),

    # Ronda 3: Despliegue final paralelo
    (None, 62, "Despliegue Railway → Render, dominio publico y SSL", "Marco", 12, 58, "Sprint"),
    (None, 63, "Optimizacion frontend: lazy loading SimilarNews, code splitting", "Luis", 11, 59, "Sprint"),
    (None, 64, "Conclusiones preliminares y recomendaciones del proyecto", "Ulises", 13, 60, "Sprint"),
    (None, 65, "Pruebas de rendimiento y carga en produccion (Render)", "Marco", 12, 61, "Sprint"),

    ("V", None, "v0.8.0 - Version desplegada en Render con dominio publico, SSL, optimizaciones frontend y pruebas de rendimiento", None, 1, 65, "Iteracion"),

    # Revision post-Ciclo 4
    (None, 89, "Revision Ciclo 4: validacion del despliegue, pruebas y rendimiento", "Todos", 5, 65, "Sprint"),

    ("V", None, "v0.9.0 - Validacion post-despliegue completada, rendimiento verificado y ajustes finales de integracion", None, 1, 89, "Iteracion"),

    # ══════════ CICLO 5: DOCUMENTACION FINAL Y CIERRE ══════════
    (None, None, None, None, None, None, None),

    # Ronda 1: Validacion con datos reales paralelo
    (None, 66, "Analisis con URLs reales: Milenio, Reforma, Aristegui, estafas", "Marco", 12, 89, "Sprint"),
    (None, 67, "Diseno responsive: ajustes layout para movil, tablet y escritorio", "Luis", 11, 89, "Sprint"),
    (None, 68, "Capitulo 5: Resultados, pruebas, analisis de datos y conclusiones", "Ulises", 13, 89, "Sprint"),
    (None, 69, "Pruebas de regresion post-despliegue y verificacion funcional", "Luis", 12, 89, "Sprint"),

    ("V", None, "v0.10.0 - Analisis con URLs reales, diseno responsive, resultados preliminares y pruebas de regresion post-despliegue", None, 1, 69, "Iteracion"),

    # Ronda 2: Correcciones y documentacion paralelo
    (None, 70, "Correccion errores: timeouts, override de dominios, edge cases URLs", "Marco", 12, 66, "Sprint"),
    (None, 71, "Manual de usuario completo y guia de uso del sistema", "Luis", 11, 67, "Sprint"),
    (None, 72, "Introduccion, resumen y abstract en espanol e ingles", "Ulises", 13, 68, "Sprint"),
    (None, 73, "Documentacion tecnica: manual API, arquitectura, despliegue", "Ulises", 12, 69, "Sprint"),

    ("V", None, "v1.0.0 - Documentacion completa, manuales, resultados consolidados y version estable del sistema", None, 1, 73, "Iteracion"),

    # Ronda 3: Revision final paralelo
    (None, 74, "Verificacion funcionalidad, pruebas de humo y validacion URLs reales", "Marco", 12, 70, "Sprint"),
    (None, 75, "Maquetacion tesis Word: estilos, indices, tablas de contenido", "Luis", 11, 71, "Sprint"),
    (None, 76, "Revision ortografia, gramatica y consistencia de la tesis", "Ulises", 13, 72, "Sprint"),
    (None, 77, "Correcciones estilo, citas y normas institucionales", "Ulises", 12, 73, "Sprint"),

    ("V", None, "v1.1.0 - Revisiones finales, verificacion funcional, maquetacion de tesis y correcciones de estilo", None, 1, 77, "Iteracion"),

    # Ronda 4: Revision cruzada paralelo
    (None, 78, "Revision tecnica tesis: datos, consistencia Capitulos 4-5 y codigo", "Marco", 18, 74, "Sprint"),
    (None, 79, "Material presentacion defensa: diapositivas, demo y resultados", "Luis", 17, 75, "Sprint"),
    (None, 80, "Correcciones finales segun retroalimentacion del asesor", "Ulises", 19, 76, "Sprint"),
    (None, 81, "Preparacion defensa tesis: diapositivas, demo en vivo y Q&A", "Luis", 18, 77, "Sprint"),

    # Ronda 5: Revision final y aprobacion paralelo
    (None, 90, "Revision y aprobacion del asesor: validacion final tesis y sistema", "Todos", 12, 78, "Sprint"),

    # Ronda 6: Cierre paralelo
    (None, 82, "Cierre proyecto: limpieza codigo, README, tag v1.2.0 en git", "Marco", 20, 90, "Sprint"),
    (None, 83, "PDF final de tesis (TESIS_VERIFEX.pdf) con formato definitivo", "Luis", 19, 90, "Sprint"),
    (None, 84, "Revision final y validacion de entrega segun requisitos", "Ulises", 21, 90, "Sprint"),
    (None, 85, "Pruebas finales de humo y validacion de cierre del proyecto", "Luis", 20, 90, "Sprint"),

    ("V", None, "v1.2.0 - Tesis completa, sistema VERIFEX version final y presentacion de defensa preparada", None, 1, 85, "Iteracion"),
]

phase_names = [
    "Fase 0: Eleccion y Preparacion del Tema",
    "Fase 1: Marco Teorico e Historias de Usuario",
    "Fase 2: Diseno del Sistema",
    "Fase 3: Desarrollo del Sistema",
    "Fase 4: Validacion y Despliegue",
    "Fase 5: Resultados y Documentacion Final",
]

cycle_names = [
    "Ciclo 0: Gestion de Tesis",
    "Ciclo 1: Fundamentos Teoricos y Requisitos",
    "Ciclo 2: Diseno del Sistema",
    "Ciclo 3: Desarrollo del Sistema",
    "Ciclo 4: Validacion y Despliegue",
    "Ciclo 5: Resultados y Documentacion Final",
]

assignee_styles = {
    "Marco": (MARCO_FILL, MARCO_FONT),
    "Luis": (LUIS_FILL, LUIS_FONT),
    "Ulises": (ULISES_FILL, ULISES_FONT),
    "Todos": (PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"),
              Font(name="Calibri", size=10, bold=True, color="616161")),
}

status_map = {"H": "Completado", "P": "En Progreso", "N": "Pendiente", "V": "Lanzado"}
status_fills = {
    "H": PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
    "P": PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid"),
    "N": PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid"),
    "V": PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"),
}
status_fonts = {
    "H": Font(name="Calibri", size=10, bold=True, color="2E7D32"),
    "P": Font(name="Calibri", size=10, bold=True, color="F57F17"),
    "N": Font(name="Calibri", size=10, bold=True, color="C62828"),
    "V": Font(name="Calibri", size=10, bold=True, color="E65100"),
}

task_dates = {}
phase_idx = -1
row = 2
version_counter = 0

cycle_fill = PatternFill(start_color="C5CAE9", end_color="C5CAE9", fill_type="solid")
tipo_font = Font(name="Calibri", size=10, bold=True, color="1565C0")

for td in task_defs:
    marker, task_num, task_name, assignee, duration, dep, tipo = td
    ncols = 10

    if marker is None and task_num is None:
        phase_idx += 1

        # Cycle header
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        cell = ws.cell(row=row, column=1, value=cycle_names[phase_idx])
        cell.font = Font(name="Calibri", bold=True, size=12, color="1A237E")
        cell.fill = cycle_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = thin_border
        for c in range(2, ncols + 1):
            ws.cell(row=row, column=c).border = thin_border
            ws.cell(row=row, column=c).fill = cycle_fill
        row += 1

        # Phase header
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        cell = ws.cell(row=row, column=1, value=phase_names[phase_idx])
        cell.font = phase_font
        cell.fill = phase_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = thin_border
        for c in range(2, ncols + 1):
            ws.cell(row=row, column=c).border = thin_border
            ws.cell(row=row, column=c).fill = phase_fill
        row += 1
        continue

    if marker == "V":
        dep_key = f"T{dep}"
        if dep_key in task_dates:
            start = task_dates[dep_key][1] + timedelta(days=1)
        else:
            start = date(2025, 9, 1)
        end = start + timedelta(days=duration)

        vkey = f"V{version_counter}"
        task_dates[vkey] = (start, end)
        version_counter += 1

        dep_str = str(dep) if dep else ""
        values = ["", tipo, "★", task_name, "—", start, end, duration + 1, dep_str, "Lanzado"]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val if val != "" else None)
            cell.font = VERSION_FONT
            cell.fill = VERSION_FILL
            cell.border = thin_border
            cell.alignment = center if col_idx in (2, 3, 5, 6, 7, 8, 9, 10) else left_wrap
            if col_idx in (6, 7):
                cell.number_format = "DD/MM/YYYY"
    else:
        key = f"T{task_num}"
        dep_key = f"T{dep}" if dep else None
        if dep_key and dep_key in task_dates:
            start = task_dates[dep_key][1] + timedelta(days=1)
        else:
            start = date(2025, 9, 1)
        end = start + timedelta(days=duration)
        task_dates[key] = (start, end)

        status = "H"

        status_text = status_map[status]
        a_fill, a_font = assignee_styles.get(assignee, (None, Font(name="Calibri", size=10)))

        dep_str = str(dep) if dep else ""
        values = ["", tipo, task_num, task_name, assignee, start, end, duration + 1, dep_str, status_text]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val if val != "" else None)
            cell.font = Font(name="Calibri", size=10)
            cell.border = thin_border
            cell.alignment = center if col_idx in (2, 3, 5, 6, 7, 8, 9, 10) else left_wrap
            if col_idx in (6, 7):
                cell.number_format = "DD/MM/YYYY"
            elif col_idx == 2:
                cell.font = tipo_font
            elif col_idx == 5:
                if a_fill:
                    cell.fill = a_fill
                cell.font = a_font
            elif col_idx == 10:
                cell.fill = status_fills.get(status)
                cell.font = status_fonts.get(status)
    row += 1

# ── Legend ──
row += 2
ws.cell(row=row, column=1, value="Leyenda:").font = Font(name="Calibri", bold=True, size=11)
row += 1
legend_items = [
    ("Marco - Backend, frontend, IA, deploy y pruebas", MARCO_FILL, MARCO_FONT),
    ("Luis - Frontend, diseno UI/UX y estilos visuales", LUIS_FILL, LUIS_FONT),
    ("Ulises - Documentacion y redaccion de tesis", ULISES_FILL, ULISES_FONT),
    ("★ Iteracion - Hito de lanzamiento de version", VERSION_FILL, VERSION_FONT),
    ("Sprint - Tarea individual con duracion definida", VERSION_FILL, VERSION_FONT),
    ("Ciclo - Gran fase del proyecto (agrupacion de fases)", VERSION_FILL, VERSION_FONT),
]
for text, fill, font in legend_items:
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = font
    cell.fill = fill
    cell.border = thin_border
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    for c in range(2, 5):
        ws.cell(row=row, column=c).border = thin_border
        ws.cell(row=row, column=c).fill = fill
    row += 1

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:J{row - 10}"
for r in range(2, row):
    ws.row_dimensions[r].height = 24
ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.page_setup.orientation = "landscape"

output_path = f"{BASE}/Gantt_VERIFEX.xlsx"
wb.save(output_path)
print(f"Excel generado: {output_path}")

# Print version timeline
print("\nLinea de versiones del sistema:")
for td in task_defs:
    if len(td) >= 7 and td[0] == "V":
        name = td[2]
        dep = td[5]
        dep_key = f"T{dep}"
        if dep_key in task_dates:
            s, e = task_dates[dep_key]
            nd = e + timedelta(days=1)
            print(f"  {name:<65} -> {nd.strftime('%d/%m/%Y')}")

# Print per-person task count
from collections import Counter
person_counts = Counter()
for td in task_defs:
    if len(td) >= 5 and td[0] is None and td[1] is not None:
        person_counts[td[3]] += 1
print("\nTareas por persona:")
for person, count in sorted(person_counts.items()):
    print(f"  {person}: {count} tareas")
