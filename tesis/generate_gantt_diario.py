#!/usr/bin/env python3
"""Generador de Gantt VERIFEX - Vista Dia a Dia
- Misma data que el Gantt principal (generate_gantt_excel.py)
- Cada columna es un dia calendario
- Cada fila es una tarea con barras de color en los dias activos
- Colores por asignado para visibilidad inmediata
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, timedelta
from collections import OrderedDict

BASE = "/Users/maarco_serrano/Downloads/verifex-standalone 2"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Gantt Diario"

# ── Colores por persona ──
MARCO_FILL = PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")
MARCO_FONT = Font(name="Calibri", size=8, bold=True, color="4A148C")
MARCO_ACTIVE = PatternFill(start_color="CE93D8", end_color="CE93D8", fill_type="solid")
TONY_FILL = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
TONY_FONT = Font(name="Calibri", size=8, bold=True, color="B71C1C")
TONY_ACTIVE = PatternFill(start_color="EF9A9A", end_color="EF9A9A", fill_type="solid")
LUIS_FILL = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
LUIS_FONT = Font(name="Calibri", size=8, bold=True, color="0D47A1")
LUIS_ACTIVE = PatternFill(start_color="90CAF9", end_color="90CAF9", fill_type="solid")
ULISES_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
ULISES_FONT = Font(name="Calibri", size=8, bold=True, color="1B5E20")
ULISES_ACTIVE = PatternFill(start_color="A5D6A7", end_color="A5D6A7", fill_type="solid")
TODOS_ACTIVE = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
VERSION_ACTIVE = PatternFill(start_color="FFE082", end_color="FFE082", fill_type="solid")

FILLS_ACTIVE = {
    "Marco": MARCO_ACTIVE,
    "Luis": LUIS_ACTIVE,
    "Ulises": ULISES_ACTIVE,
    "Tony": TONY_ACTIVE,
    "Todos": TODOS_ACTIVE,
}

assignee_styles = {
    "Marco": (MARCO_FILL, MARCO_FONT),
    "Luis": (LUIS_FILL, LUIS_FONT),
    "Ulises": (ULISES_FILL, ULISES_FONT),
    "Tony": (TONY_FILL, TONY_FONT),
    "Todos": (PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"),
              Font(name="Calibri", size=8, bold=True, color="616161")),
}

thin_border = Border(
    left=Side(style="thin", color="BDBDBD"),
    right=Side(style="thin", color="BDBDBD"),
    top=Side(style="thin", color="BDBDBD"),
    bottom=Side(style="thin", color="BDBDBD"),
)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=8)
header_fill = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
phase_font = Font(name="Calibri", bold=True, color="FFFFFF", size=9)
phase_fill = PatternFill(start_color="3949AB", end_color="3949AB", fill_type="solid")
cycle_fill = PatternFill(start_color="C5CAE9", end_color="C5CAE9", fill_type="solid")
cycle_font = Font(name="Calibri", bold=True, size=10, color="1A237E")
tipo_font = Font(name="Calibri", size=8, bold=True, color="1565C0")
VERSION_FONT = Font(name="Calibri", size=8, bold=True, color="E65100")

# ── Task definitions (identical to generate_gantt_excel.py) ──
task_defs = [
    (None, None, None, None, None, None, None),
(None, 1, "Eleccion del tema: deteccion de noticias falsas con IA", "Todos", 9, None, "Sprint"),
(None, 2, "Investigacion preliminar y revision bibliografica (fake news, IA)", "Todos", 22, 1, "Sprint"),
(None, 3, "Definicion del problema, preguntas de investigacion y objetivos", "Todos", 5, 2, "Sprint"),
(None, 4, "Aprobacion formal del tema y delimitacion del alcance", "Todos", 3, 3, "Sprint"),
    (None, None, None, None, None, None, None),
(None, 5, "Backend e IA: Groq API, Flask, scraping, curl_cffi, Playwright", "Marco", 18, 4, "Sprint"),
(None, 6, "UI/UX cyberpunk: React, Tailwind, tipografia, paletas", "Luis", 17, 4, "Sprint"),
(None, 7, "Marco teorico: fake news, verificacion digital, desinformacion MX", "Ulises", 19, 4, "Sprint"),
(None, 8, "Herramientas de prueba: Vitest, Pytest, Playwright, Selenium", "Tony", 18, 4, "Sprint"),
(None, 9, "Seleccion de tecnologias y definicion de requisitos funcionales", "Marco", 12, 5, "Sprint"),
(None, 10, "Historias de usuario (HU-01 a HU-10) y diseno conceptual de interfaz", "Luis", 11, 6, "Sprint"),
(None, 11, "Planteamiento del problema, justificacion y preguntas de investigacion", "Ulises", 13, 7, "Sprint"),
(None, 12, "Metricas de calidad, criterios de aceptacion y matriz de trazabilidad", "Tony", 12, 8, "Sprint"),
(None, 13, "Setup de entorno: Node, Python, Git, VSCode, dependencias", "Marco", 10, 9, "Sprint"),
(None, 14, "Moodboard, exploracion de componentes UI y paleta de colores", "Luis", 9, 10, "Sprint"),
(None, 15, "Capitulo 1: Introduccion (contexto, problema, objetivos, alcance)", "Ulises", 11, 11, "Sprint"),
(None, 16, "Plan maestro de pruebas y diseno de casos de prueba iniciales", "Tony", 10, 12, "Sprint"),
(None, 17, "Revision consolidacion: requisitos, HU y plan de pruebas", "Todos", 7, 16, "Sprint"),
    ("V", None, "v0.1.0 - Requisitos, HU, marco teorico y plan de pruebas definidos", None, 1, 17, "Iteracion"),

    # Revision post-Ciclo 1
    (None, 86, "Revision y retroalimentacion del Ciclo 1: ajustes a requisitos, HU y marco teorico", "Todos", 5, 17, "Sprint"),

    (None, None, None, None, None, None, None),
(None, 18, "Arquitectura cliente-servidor: React + Flask + Groq API + Playwright + CORS", "Marco", 12, 86, "Sprint"),
(None, 19, "Wireframes baja fidelidad y prototipo navegable alta fidelidad", "Luis", 11, 86, "Sprint"),
(None, 20, "Capitulo 2: Marco Teorico (fake news, IA, Groq API, verificacion)", "Ulises", 13, 86, "Sprint"),
(None, 21, "Diagramas UML: casos de uso, actividades, secuencia, clases, ER", "Tony", 12, 86, "Sprint"),
(None, 22, "Setup proyecto: Vite + React + TypeScript + Tailwind + PostCSS", "Marco", 8, 18, "Sprint"),
(None, 23, "Paleta colores definitiva, tipografia (Orbitron, Rajdhani) y guia de estilos", "Luis", 7, 19, "Sprint"),
(None, 24, "Diagrama Gantt y tablero Kanban (Gantt_VERIFEX.xlsx, Kanban_VERIFEX.xlsx)", "Ulises", 9, 20, "Sprint"),
(None, 25, "Datos de prueba y configuracion del entorno de validacion", "Tony", 8, 21, "Sprint"),
(None, 26, "Diseno detallado del scraper multi-estrategia y plan de extraccion", "Marco", 10, 22, "Sprint"),
(None, 27, "Maquetacion: UrlInput, layout inicial y estructura de paneles", "Luis", 9, 23, "Sprint"),
(None, 28, "Capitulo 3: Metodologia y Diseno del Sistema (arquitectura, metodos)", "Ulises", 11, 24, "Sprint"),
(None, 29, "Casos de prueba detallados para scraper y extraccion de contenido", "Tony", 10, 25, "Sprint"),
    ("V", None, "v0.2.0 - Diseno completo: arquitectura, prototipo navegable, diagramas UML y plan de pruebas detallado", None, 1, 29, "Iteracion"),

    # Revision post-Ciclo 2
    (None, 87, "Revision y retroalimentacion del Ciclo 2: validacion de diseno, prototipo y plan de pruebas", "Todos", 5, 29, "Sprint"),

    (None, None, None, None, None, None, None),
(None, 30, "Setup Flask: rutas /analyze y /health, CORS, parse_response, errores", "Marco", 8, 87, "Sprint"),
(None, 31, "Componentes base: VerdictDisplay, ConfidenceBar, veredicto por color", "Luis", 7, 87, "Sprint"),
(None, 32, "Resultados esperados, metricas de evaluacion y criterios de exito", "Ulises", 9, 87, "Sprint"),
(None, 33, "Configuracion de pruebas: pytest, vitest, jsdom, testing-library", "Tony", 8, 87, "Sprint"),
(None, 34, "Scraper multi-estrategia: cloudscraper, curl_cffi, requests, Playwright", "Marco", 12, 30, "Sprint"),
(None, 35, "Componentes: RedFlags, SimilarNews, LanguageToggle con estilos", "Luis", 11, 31, "Sprint"),
(None, 36, "Documentacion de avances del Capitulo 4: Desarrollo e Implementacion", "Ulises", 13, 32, "Sprint"),
(None, 37, "Pruebas del scraper y validacion de extraccion de contenido HTML", "Tony", 12, 33, "Sprint"),
    ("V", None, "v0.3.0 - Scraper multi-estrategia funcional: cloudscraper, curl_cffi, requests y Playwright integrados y validados", None, 1, 37, "Iteracion"),
(None, 38, "Llamadas a Groq API: call_groq, prompt engineering, parse_response", "Marco", 12, 34, "Sprint"),
(None, 39, "Manejo de estados (loading, error, results) y soporte bilingue ES/EN", "Luis", 11, 35, "Sprint"),
(None, 40, "Actualizacion Capitulos 1-3 segun cambios de implementacion", "Ulises", 13, 36, "Sprint"),
(None, 41, "Pruebas de integracion scraper + Groq API con distintos tipos de URL", "Tony", 12, 37, "Sprint"),
(None, 42, "Clasificador de credibilidad: 5 categorias, CREDIBLE_DOMAINS y override", "Marco", 12, 38, "Sprint"),
(None, 43, "Diseno visual cyberpunk: cuadricula, glitch, vignette CRT, scanlines", "Luis", 11, 39, "Sprint"),
(None, 44, "Tabla de resultados, analisis preliminar de datos e indice de tesis", "Ulises", 13, 40, "Sprint"),
(None, 45, "Pruebas de integracion frontend-backend con datos simulados y reales", "Tony", 12, 41, "Sprint"),
    ("V", None, "v0.4.0 - Clasificador de credibilidad, integracion frontend-backend y diseno cyberpunk completos", None, 1, 45, "Iteracion"),
(None, 46, "Busqueda de noticias similares: news_finder.py, Google News RSS", "Marco", 12, 42, "Sprint"),
(None, 47, "Conexion frontend-backend via API REST (fetch /analyze + AbortController)", "Luis", 11, 43, "Sprint"),
(None, 48, "Referencias bibliograficas, anexos y plantilla oficial de tesis", "Ulises", 13, 44, "Sprint"),
(None, 49, "Pruebas de regresion y deteccion de defectos en integracion continua", "Tony", 12, 45, "Sprint"),
(None, 50, "Article_type: 5 tipos y ajustes finales de clasificacion", "Marco", 10, 46, "Sprint"),
(None, 51, "Manejo de errores de red, timeouts y UX de carga animada", "Luis", 9, 47, "Sprint"),
(None, 52, "Resultados parciales, graficas y tablas de visualizacion", "Ulises", 11, 48, "Sprint"),
(None, 53, "Pruebas de aceptacion de usuario (UAT) con escenarios reales", "Tony", 10, 49, "Sprint"),
    ("V", None, "v0.5.0 - Desarrollo completo: scraper, IA, clasificador, news_finder, UI cyberpunk y pruebas de integracion", None, 1, 53, "Iteracion"),

    # Revision post-Ciclo 3
    (None, 88, "Revision Ciclo 3: validacion del desarrollo completo y correcciones", "Todos", 5, 53, "Sprint"),

    (None, None, None, None, None, None, None),
(None, 54, "Suite pruebas backend: 27 tests en server/test_analyzer.py (pytest)", "Marco", 12, 88, "Sprint"),
(None, 55, "Suite pruebas frontend: 52 tests con vitest + testing-library", "Luis", 11, 88, "Sprint"),
(None, 56, "Resultados detallados y analisis de datos recopilados", "Ulises", 13, 88, "Sprint"),
(None, 57, "Documentacion de pruebas: defectos, cobertura y metricas", "Tony", 12, 88, "Sprint"),
    ("V", None, "v0.6.0 - Suites completas de pruebas backend y frontend, resultados documentados y metricas de cobertura", None, 1, 57, "Iteracion"),
(None, 58, "Configuracion despliegue: Procfile, build.sh, CORS, gunicorn, Railway", "Marco", 12, 54, "Sprint"),
(None, 59, "Capturas del sistema: interfaz, resultados, errores, responsive", "Luis", 11, 55, "Sprint"),
(None, 60, "Actualizacion bibliografia, graficas y tablas de visualizacion", "Ulises", 13, 56, "Sprint"),
(None, 61, "Validacion de seguridad: manejo errores, proteccion API key, estres", "Tony", 12, 57, "Sprint"),
    ("V", None, "v0.7.0 - Configuracion de despliegue, capturas del sistema, validacion de seguridad y bibliografia actualizada", None, 1, 61, "Iteracion"),
(None, 62, "Despliegue Railway → Render, dominio publico y SSL", "Marco", 12, 58, "Sprint"),
(None, 63, "Optimizacion frontend: lazy loading SimilarNews, code splitting", "Luis", 11, 59, "Sprint"),
(None, 64, "Conclusiones preliminares y recomendaciones del proyecto", "Ulises", 13, 60, "Sprint"),
(None, 65, "Pruebas de rendimiento y carga en produccion (Render)", "Tony", 12, 61, "Sprint"),
    ("V", None, "v0.8.0 - Version desplegada en Render con dominio publico, SSL, optimizaciones frontend y pruebas de rendimiento", None, 1, 65, "Iteracion"),

    # Revision post-Ciclo 4
    (None, 89, "Revision Ciclo 4: validacion del despliegue, pruebas y rendimiento", "Todos", 5, 65, "Sprint"),

    ("V", None, "v0.9.0 - Validacion post-despliegue completada, rendimiento verificado y ajustes finales de integracion", None, 1, 89, "Iteracion"),

    # ══════════ CICLO 5: DOCUMENTACION FINAL Y CIERRE ══════════
    (None, None, None, None, None, None, None),

    # Ronda 1: Validacion con datos reales paralelo
(None, 66, "Analisis con URLs reales: Milenio, Reforma, Aristegui, estafas", "Marco", 12, 89, "Sprint"),
(None, 67, "Diseno responsive: ajustes layout para movil, tablet y escritorio", "Luis", 11, 89, "Sprint"),
(None, 68, "Capitulo 5: Resultados, pruebas, analisis de datos y conclusiones", "Ulises", 13, 89, "Sprint"),
(None, 69, "Pruebas de regresion post-despliegue y verificacion funcional", "Tony", 12, 89, "Sprint"),

    ("V", None, "v0.10.0 - Analisis con URLs reales, diseno responsive, resultados preliminares y pruebas de regresion post-despliegue", None, 1, 69, "Iteracion"),

    # Ronda 2: Correcciones y documentacion paralelo
(None, 70, "Correccion errores: timeouts, override de dominios, edge cases URLs", "Marco", 12, 66, "Sprint"),
(None, 71, "Manual de usuario completo y guia de uso del sistema", "Luis", 11, 67, "Sprint"),
(None, 72, "Introduccion, resumen y abstract en espanol e ingles", "Ulises", 13, 68, "Sprint"),
(None, 73, "Documentacion tecnica: manual API, arquitectura, despliegue", "Tony", 12, 69, "Sprint"),
    ("V", None, "v1.0.0 - Documentacion completa, manuales, resultados consolidados y version estable del sistema", None, 1, 73, "Iteracion"),
    # Ronda 3: Revision final paralelo
(None, 74, "Verificacion funcionalidad, pruebas de humo y validacion URLs reales", "Marco", 12, 70, "Sprint"),
(None, 75, "Maquetacion tesis Word: estilos, indices, tablas de contenido", "Luis", 11, 71, "Sprint"),
(None, 76, "Revision ortografia, gramatica y consistencia de la tesis", "Ulises", 13, 72, "Sprint"),
(None, 77, "Correcciones estilo, citas y normas institucionales", "Tony", 12, 73, "Sprint"),
    ("V", None, "v1.1.0 - Revisiones finales, verificacion funcional, maquetacion de tesis y correcciones de estilo", None, 1, 77, "Iteracion"),

    # Ronda 4: Revision cruzada paralelo
(None, 78, "Revision tecnica tesis: datos, consistencia Capitulos 4-5 y codigo", "Marco", 18, 74, "Sprint"),
(None, 79, "Material presentacion defensa: diapositivas, demo y resultados", "Luis", 17, 75, "Sprint"),
(None, 80, "Correcciones finales segun retroalimentacion del asesor", "Ulises", 19, 76, "Sprint"),
(None, 81, "Preparacion defensa tesis: diapositivas, demo en vivo y Q&A", "Tony", 18, 77, "Sprint"),

    # Ronda 5: Revision final y aprobacion paralelo
    (None, 90, "Revision y aprobacion del asesor: validacion final tesis y sistema", "Todos", 12, 78, "Sprint"),

    # Ronda 6: Cierre paralelo
(None, 82, "Cierre proyecto: limpieza codigo, README, tag v1.2.0 en git", "Marco", 20, 90, "Sprint"),
(None, 83, "PDF final de tesis (TESIS_VERIFEX.pdf) con formato definitivo", "Luis", 19, 90, "Sprint"),
(None, 84, "Revision final y validacion de entrega segun requisitos", "Ulises", 21, 90, "Sprint"),
(None, 85, "Pruebas finales de humo y validacion de cierre del proyecto", "Tony", 20, 90, "Sprint"),

    ("V", None, "v1.2.0 - Tesis completa, sistema VERIFEX version final y presentacion de defensa preparada", None, 1, 85, "Iteracion"),
]

phase_names = [
    "Fase 0: Eleccion y Preparacion del Tema",
    "Fase 1: Marco Teorico e Historias de Usuario",
    "Fase 2: Diseno del Sistema",
    "Fase 3: Desarrollo del Sistema",
    "Fase 4: Validacion y Despliegue",
    "Fase 5: Resultados y Documentacion Final",
]

cycle_names = [
    "Ciclo 0: Gestion de Tesis",
    "Ciclo 1: Fundamentos Teoricos y Requisitos",
    "Ciclo 2: Diseno del Sistema",
    "Ciclo 3: Desarrollo del Sistema",
    "Ciclo 4: Validacion y Despliegue",
    "Ciclo 5: Resultados y Documentacion Final",
]

# ── Calcular fechas de cada tarea (misma logica que el Gantt principal) ──
task_dates = {}
version_counter = 0
for td in task_defs:
    marker, task_num, task_name, assignee, duration, dep, tipo = td
    if marker is None and task_num is None:
        continue
    if marker == "V":
        dep_key = f"T{dep}"
        if dep_key in task_dates:
            start = task_dates[dep_key][1] + timedelta(days=1)
        else:
            start = date(2025, 9, 1)
        end = start + timedelta(days=duration)
        task_dates[f"V{version_counter}"] = (start, end)
        version_counter += 1
    else:
        key = f"T{task_num}"
        dep_key = f"T{dep}" if dep else None
        if dep_key and dep_key in task_dates:
            start = task_dates[dep_key][1] + timedelta(days=1)
        else:
            start = date(2025, 9, 1)
        end = start + timedelta(days=duration)
        task_dates[key] = (start, end)

# ── Determinar rango de fechas ──
all_dates_flattened = [d for dates in task_dates.values() for d in dates]
start_date = min(all_dates_flattened)
end_date = max(all_dates_flattened)

days = []
current = start_date
while current <= end_date:
    days.append(current)
    current += timedelta(days=1)

INFO_COLS = 4
DATE_START_COL = INFO_COLS + 1

# ── Establecer anchos de columna ──
ws.column_dimensions["A"].width = 4
ws.column_dimensions["B"].width = 72
ws.column_dimensions["C"].width = 10
ws.column_dimensions["D"].width = 10
for i in range(len(days)):
    col_letter = get_column_letter(DATE_START_COL + i)
    ws.column_dimensions[col_letter].width = 2.5

# ── Agrupar meses ──
month_ranges = OrderedDict()
for d in days:
    key = d.strftime("%b %Y")
    if key not in month_ranges:
        month_ranges[key] = {"start": d, "end": d}
    else:
        month_ranges[key]["end"] = d

# ── Fila 1: Encabezados de mes ──
WE_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

# Escribir encabezados de info
info_headers = ["#", "Tarea", "Asignado", "Tipo"]
for ci, h in enumerate(info_headers, 1):
    cell = ws.cell(row=1, column=ci, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = thin_border

# Encabezados de mes
for month_info in month_ranges.values():
    start_col = DATE_START_COL + days.index(month_info["start"])
    end_col = DATE_START_COL + days.index(month_info["end"])
    ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
    cell = ws.cell(row=1, column=start_col, value=month_info["start"].strftime("%b %Y"))
    cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=6)
    cell.fill = PatternFill(start_color="3949AB", end_color="3949AB", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
    for c in range(start_col, end_col + 1):
        ws.cell(row=1, column=c).border = thin_border
        ws.cell(row=1, column=c).fill = PatternFill(start_color="3949AB", end_color="3949AB", fill_type="solid")

# ── Fila 2: Numeros de dia ──
for i, d in enumerate(days):
    col = DATE_START_COL + i
    cell = ws.cell(row=2, column=col, value=d.day)
    cell.font = Font(name="Calibri", size=5, color="666666")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border
    if d.weekday() in (5, 6):
        cell.fill = WE_FILL
    else:
        cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

# ── Rellenar tareas ──
row = 3
phase_idx = -1
version_counter = 0

for td in task_defs:
    marker, task_num, task_name, assignee, duration, dep, tipo = td
    ncols_total = INFO_COLS + len(days)

    if marker is None and task_num is None:
        phase_idx += 1
        # Ciclo header
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols_total)
        cell = ws.cell(row=row, column=1, value=cycle_names[phase_idx])
        cell.font = cycle_font
        cell.fill = cycle_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = thin_border
        for c in range(2, ncols_total + 1):
            ws.cell(row=row, column=c).border = thin_border
            ws.cell(row=row, column=c).fill = cycle_fill
        ws.row_dimensions[row].height = 20
        row += 1

        # Fase header
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols_total)
        cell = ws.cell(row=row, column=1, value=phase_names[phase_idx])
        cell.font = phase_font
        cell.fill = phase_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = thin_border
        for c in range(2, ncols_total + 1):
            ws.cell(row=row, column=c).border = thin_border
            ws.cell(row=row, column=c).fill = phase_fill
        ws.row_dimensions[row].height = 18
        row += 1
        continue

    # Obtener fechas de esta tarea
    if marker == "V":
        key = f"V{version_counter}"
        version_counter += 1
    else:
        key = f"T{task_num}"

    if key not in task_dates:
        row += 1
        continue
    t_start, t_end = task_dates[key]

    # Escribir columnas de informacion
    is_version = marker == "V"
    if is_version:
        ws.cell(row=row, column=1, value="★").font = VERSION_FONT
        ws.cell(row=row, column=2, value=task_name).font = VERSION_FONT
        ws.cell(row=row, column=3, value="—").font = VERSION_FONT
        ws.cell(row=row, column=4, value="Iteracion").font = VERSION_FONT
        active_fill = VERSION_ACTIVE
    else:
        c1 = ws.cell(row=row, column=1, value=task_num)
        c1.font = Font(name="Calibri", size=8)
        c2 = ws.cell(row=row, column=2, value=task_name)
        c2.font = Font(name="Calibri", size=8)
        a_fill, a_font = assignee_styles.get(assignee, (None, Font(name="Calibri", size=8)))
        c3 = ws.cell(row=row, column=3, value=assignee if assignee else "")
        c3.font = a_font
        c3.fill = a_fill if a_fill else PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        c4 = ws.cell(row=row, column=4, value=tipo)
        c4.font = tipo_font
        active_fill = FILLS_ACTIVE.get(assignee, FILLS_ACTIVE["Todos"])

    for ci in range(1, INFO_COLS + 1):
        cell = ws.cell(row=row, column=ci)
        cell.border = thin_border
        cell.alignment = left_wrap if ci == 2 else center

    # Rellenar celdas de fecha
    for i, d in enumerate(days):
        col = DATE_START_COL + i
        cell = ws.cell(row=row, column=col)
        cell.border = thin_border
        if t_start <= d <= t_end:
            cell.fill = active_fill
        elif d.weekday() in (5, 6):
            cell.fill = WE_FILL
        else:
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    ws.row_dimensions[row].height = 16 if is_version else 18
    row += 1

# ── Leyenda ──
row += 2
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
ws.cell(row=row, column=1, value="Leyenda:").font = Font(name="Calibri", bold=True, size=10)
row += 1
legend_items = [
    ("Marco - Backend, frontend, IA, deploy y pruebas", MARCO_ACTIVE),
    ("Luis - Frontend, diseno UI/UX y estilos visuales", LUIS_ACTIVE),
    ("Ulises - Documentacion y redaccion de tesis", ULISES_ACTIVE),
    ("Tony - Diagramas UML, documentacion y revision de logica", TONY_ACTIVE),
    ("★ Iteracion - Hito de lanzamiento de version", VERSION_ACTIVE),
]
for text, fill in legend_items:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(name="Calibri", size=9, bold=True)
    cell.fill = fill
    cell.border = thin_border
    for c in range(2, 5):
        ws.cell(row=row, column=c).border = thin_border
        ws.cell(row=row, column=c).fill = fill
    row += 1

ws.freeze_panes = "E3"
ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.page_setup.orientation = "landscape"

output_path = f"{BASE}/Gantt_VERIFEX_DIARIO.xlsx"
wb.save(output_path)
print(f"Gantt diario generado: {output_path}")
print(f"Rango: {start_date} -> {end_date} ({len(days)} dias)")
print(f"Tareas: {sum(1 for td in task_defs if td[0] != 'V' and td[1] is not None)} sprints")
print(f"Versiones: {version_counter} hitos")

# Per-person stats
from collections import Counter
person_counts = Counter()
for td in task_defs:
    if len(td) >= 5 and td[0] is None and td[1] is not None:
        person_counts[td[3]] += 1
print("\nTareas por persona:")
for person, count in sorted(person_counts.items()):
    print(f"  {person}: {count} tareas")
